import openpyxl
from openpyxl import load_workbook

# Deep analysis of template structure
wb = load_workbook('backend/teamp/Templat-Report.xlsx')

# Analyze Page (4) and Page (5) - the report sheets
for sheet_name in ['Page (4)', 'Page (5)']:
    sheet = wb[sheet_name]
    print(f'\n=== {sheet_name} - Full Structure ===')
    print(f'Max row: {sheet.max_row}, Max col: {sheet.max_column}')
    
    # Look for all text content
    content = []
    for row in range(1, min(60, sheet.max_row + 1)):
        for col in range(1, 15):
            cell = sheet.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str):
                content.append((row, col, cell.value))
    
    print('\nKey content:')
    for row, col, val in content[:30]:
        print(f'  [{row},{col}] {val[:60]}')

# Check merged cells
print('\n=== Merged Cells in Page (4) ===')
sheet = wb['Page (4)']
for merged_range in sheet.merged_cells.ranges:
    print(f'  {merged_range}')

wb.close()

import openpyxl
from openpyxl import load_workbook

# Load template and analyze Page (4) - the report sheet
wb = load_workbook('backend/teamp/Templat-Report.xlsx')

print('=== Page (4) Structure ===')
sheet = wb['Page (4)']

# Find photo areas and labels
print('Looking for photo areas and labels...')
for row in range(1, sheet.max_row + 1):
    for col in range(1, 15):  # Check first 14 columns
        cell = sheet.cell(row=row, column=col)
        if cell.value and isinstance(cell.value, str):
            val = cell.value.lower()
            if any(keyword in val for keyword in ['photo', 'picture', 'image', 'รูป', 'before', 'after', 'ลายเซ็น', 'signature', 'checked', 'ผู้ตรวจ']):
                print(f'  [{row},{col}] "{cell.value}"')

print()
print('=== Page (5) Structure ===')
sheet = wb['Page (5)']
for row in range(1, sheet.max_row + 1):
    for col in range(1, 15):
        cell = sheet.cell(row=row, column=col)
        if cell.value and isinstance(cell.value, str):
            val = cell.value.lower()
            if any(keyword in val for keyword in ['photo', 'picture', 'image', 'รูป', 'before', 'after', 'ลายเซ็น', 'signature', 'checked', 'ผู้ตรวจ', 'check', 'pass', 'fail', 'ผ่าน', 'ไม่ผ่าน']):
                print(f'  [{row},{col}] "{cell.value}"')

wb.close()

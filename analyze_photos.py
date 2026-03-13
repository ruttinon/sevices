import openpyxl
from openpyxl import load_workbook

# Deep analysis of photo areas in template
wb = load_workbook('backend/teamp/Templat-Report.xlsx')

# Analyze Page (4) - Photo areas
sheet = wb['Page (4)']
print('=== Page (4) - Photo Area Analysis ===')

# Look at rows 33-55 (photo section)
print('\nPhoto section content:')
for row in range(30, 56):
    row_content = []
    for col in range(1, 15):
        cell = sheet.cell(row=row, column=col)
        if cell.value and isinstance(cell.value, str):
            row_content.append(f'[{col}] {cell.value[:40]}')
    if row_content:
        print(f'Row {row}: {row_content}')

# Check for specific photo labels
print('\n=== Photo Labels Found ===')
for row in range(30, 56):
    for col in range(1, 15):
        cell = sheet.cell(row=row, column=col)
        if cell.value and isinstance(cell.value, str):
            val = cell.value
            if any(k in val for k in ['รูป', 'ภาพ', 'หัวข้อที่2', 'กวดขัน', 'เทอร์มินอล', 'ความสะอาด']):
                print(f'  [{row},{col}] "{val}"')

# Check row heights (photo areas are usually taller)
print('\n=== Row Heights (potential photo areas) ===')
for row_idx in range(30, 56):
    row = sheet.row_dimensions[row_idx]
    if row.height and row.height > 30:  # Taller rows likely photo areas
        print(f'  Row {row_idx}: height={row.height}')

wb.close()

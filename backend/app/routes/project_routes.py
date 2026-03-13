from datetime import datetime
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from sqlalchemy import func
from sqlalchemy.orm import Session

from ..config import DEFAULT_PROJECT_WORKBOOK_PATH, DEFAULT_REPORT_TEMPLATE_PATH, TEMPLATES_DIR, UPLOADS_DIR
from ..database import _relocate_project_uploads, get_db
from ..models import Customer, Meter, Panel, Project, Report, ServiceJob
from ..schemas import asset_schema, customer_schema, project_schema
from ..services.file_upload import save_upload
from ..services.excel_import import import_assets_from_project_workbook
from ..services.template_analyzer import analyze_template_file, extract_template_photo_captions, extract_template_checklist_items
from ..utils.helpers import build_public_file_path
from ..utils.project_files import ensure_project_upload_tree, project_upload_dir

router = APIRouter(tags=["customers", "projects", "dashboard"])


def _absolute_path_from_public(public_path: str):
    relative = public_path.removeprefix("/uploads/")
    return UPLOADS_DIR / relative


@router.get("/customers", response_model=list[customer_schema.CustomerResponse])
def read_customers(db: Session = Depends(get_db)):
    return db.query(Customer).order_by(Customer.name.asc()).all()


@router.post("/customers", response_model=customer_schema.CustomerResponse)
def create_customer(customer: customer_schema.CustomerCreate, db: Session = Depends(get_db)):
    db_customer = Customer(**customer.model_dump())
    db.add(db_customer)
    db.commit()
    db.refresh(db_customer)
    return db_customer


@router.put("/customers/{customer_id}", response_model=customer_schema.CustomerResponse)
def update_customer(
    customer_id: int,
    customer_update: customer_schema.CustomerUpdate,
    db: Session = Depends(get_db),
):
    customer = db.get(Customer, customer_id)
    if customer is None:
        raise HTTPException(status_code=404, detail="Customer not found")

    for field, value in customer_update.model_dump(exclude_unset=True).items():
        setattr(customer, field, value)

    db.commit()
    _relocate_project_uploads()
    db.refresh(customer)
    return customer


@router.get("/projects", response_model=list[project_schema.ProjectResponse])
def read_projects(db: Session = Depends(get_db)):
    return db.query(Project).order_by(Project.created_at.desc()).all()


@router.post("/projects", response_model=project_schema.ProjectResponse)
def create_project(project: project_schema.ProjectCreate, db: Session = Depends(get_db)):
    customer = db.get(Customer, project.customer_id)
    if customer is None:
        raise HTTPException(status_code=404, detail="Customer not found")

    db_project = Project(**project.model_dump())
    db.add(db_project)
    db.commit()
    db.refresh(db_project)
    ensure_project_upload_tree(UPLOADS_DIR, db_project)
    return db_project


@router.get("/projects/{project_id}/latest-report")
def get_latest_project_report(project_id: int, db: Session = Depends(get_db)):
    """Get the latest overall project report."""
    report = db.query(Report).filter(
        Report.project_id == project_id,
        Report.service_id == None  # Overall project reports don't have a service_id
    ).order_by(Report.created_at.desc()).first()

    if not report:
        # Try to find any report for this project if no "overall" report exists
        report = db.query(Report).filter(
            Report.project_id == project_id
        ).order_by(Report.created_at.desc()).first()

    if not report:
        return {"id": None, "file_path": None}

    return {
        "id": report.id,
        "file_path": report.file_path,
        "created_at": report.created_at,
        "report_date": report.report_date
    }


@router.get("/projects/{project_id}/public")
def get_project_public(project_id: int, db: Session = Depends(get_db)):
    project = db.get(Project, project_id)
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    # Get stats for the dashboard
    total_meters = db.query(Meter).join(Loop).join(Panel).filter(Panel.project_id == project_id).count()
    completed_meters = db.query(ServiceJob).filter(
        ServiceJob.project_id == project_id,
        ServiceJob.status == "Completed",
        ServiceJob.meter_id != None
    ).count()

    latest_report = db.query(Report).filter(Report.project_id == project_id).order_by(Report.created_at.desc()).first()

    # Include latest report files for customer overview
    report_records = (
        db.query(Report)
        .filter(Report.project_id == project_id)
        .order_by(Report.created_at.desc())
        .limit(20)
        .all()
    )

    return {
        "project": {
            "id": project.id,
            "name": project.name,
            "location": project.location,
            "customer_name": project.customer.name,
        },
        "stats": {
            "total_panels": len(project.panels),
            "total_meters": total_meters,
            "completed_meters": completed_meters,
            "latest_report_id": latest_report.id if latest_report else None,
        },
        "panels": [
            {
                "id": panel.id,
                "panel_code": panel.panel_code,
                "panel_name": panel.panel_name,
                "loops": [
                    {
                        "id": loop.id,
                        "loop_code": loop.loop_code,
                        "loop_name": loop.loop_name,
                        "meters": [
                            {
                                "id": meter.id,
                                "meter_code": meter.meter_code,
                                "meter_name": meter.meter_name,
                                "status": db.query(ServiceJob).filter(
                                    ServiceJob.meter_id == meter.id
                                ).order_by(ServiceJob.created_at.desc()).first().status if db.query(ServiceJob).filter(
                                    ServiceJob.meter_id == meter.id
                                ).first() else "Pending"
                            }
                            for meter in loop.meters
                        ]
                    }
                    for loop in panel.loops
                ]
            }
            for panel in project.panels
        ],
        "reports": [
            {
                "id": report.id,
                "file_path": report.file_path,
                "created_at": report.created_at,
                "report_date": getattr(report, 'report_date', None),
            }
            for report in report_records
        ],
    }


@router.put("/projects/{project_id}", response_model=project_schema.ProjectResponse)
def update_project(
    project_id: int,
    project_update: project_schema.ProjectUpdate,
    db: Session = Depends(get_db),
):
    project = db.get(Project, project_id)
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    update_data = project_update.model_dump(exclude_unset=True)
    if "customer_id" in update_data:
        customer = db.get(Customer, update_data["customer_id"])
        if customer is None:
            raise HTTPException(status_code=404, detail="Customer not found")

    for field, value in update_data.items():
        setattr(project, field, value)

    db.commit()
    _relocate_project_uploads()
    db.refresh(project)
    return project


@router.post("/projects/{project_id}/template", response_model=project_schema.ProjectResponse)
def upload_project_template(
    project_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    project = db.get(Project, project_id)
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    if not (file.filename or "").lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm", ".xls")):
        raise HTTPException(status_code=400, detail="Please upload an Excel report template file.")

    project_template_dir = project_upload_dir(UPLOADS_DIR, project, "documents", "templates")
    project.template_file_path = save_upload(file, project_template_dir, prefix="TPL", keep_original_name=True)
    db.commit()
    db.refresh(project)
    return project


@router.post("/projects/{project_id}/workbook", response_model=project_schema.ProjectResponse)
def upload_project_workbook(
    project_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    project = db.get(Project, project_id)
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    if not (file.filename or "").lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm", ".xls")):
        raise HTTPException(status_code=400, detail="Please upload an Excel project workbook file.")

    project_workbook_dir = project_upload_dir(UPLOADS_DIR, project, "documents", "workbooks")
    project.project_workbook_file_path = save_upload(file, project_workbook_dir, prefix="WB", keep_original_name=True)
    db.commit()
    db.refresh(project)
    return project


@router.get("/projects/{project_id}/template-analysis")
def read_project_template_analysis(project_id: int, db: Session = Depends(get_db)):
    project = db.get(Project, project_id)
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    template_path = None
    template_source = "default"
    if project.template_file_path:
        candidate_path = _absolute_path_from_public(project.template_file_path)
        if candidate_path.exists():
            template_path = candidate_path
            template_source = "project"
    if template_path is None and DEFAULT_REPORT_TEMPLATE_PATH.exists():
        template_path = DEFAULT_REPORT_TEMPLATE_PATH

    if template_path is None or not template_path.exists():
        raise HTTPException(status_code=404, detail="No Excel template available for this project")

    return {
        "project": {
            "id": project.id,
            "name": project.name,
            "location": project.location,
            "customer_name": project.customer.name,
            "template_file_path": project.template_file_path,
            "project_workbook_file_path": project.project_workbook_file_path,
        },
        "template_source": template_source,
        "analysis": analyze_template_file(template_path),
    }


@router.get("/projects/{project_id}/workbook-analysis")
def read_project_workbook_analysis(project_id: int, db: Session = Depends(get_db)):
    project = db.get(Project, project_id)
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    workbook_path = None
    workbook_source = "default"
    if project.project_workbook_file_path:
        candidate_path = _absolute_path_from_public(project.project_workbook_file_path)
        if candidate_path.exists():
            workbook_path = candidate_path
            workbook_source = "project"
    if workbook_path is None and DEFAULT_PROJECT_WORKBOOK_PATH.exists():
        workbook_path = DEFAULT_PROJECT_WORKBOOK_PATH

    if workbook_path is None or not workbook_path.exists():
        raise HTTPException(status_code=404, detail="No Excel project workbook available for this project")

    return {
        "project": {
            "id": project.id,
            "name": project.name,
            "location": project.location,
            "customer_name": project.customer.name,
            "template_file_path": project.template_file_path,
            "project_workbook_file_path": project.project_workbook_file_path,
        },
        "workbook_source": workbook_source,
        "analysis": analyze_template_file(workbook_path),
    }


@router.post("/projects/{project_id}/sync-workbook-assets", response_model=asset_schema.AssetImportResult)
def sync_project_workbook_assets(project_id: int, db: Session = Depends(get_db)):
    project = db.get(Project, project_id)
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")
    try:
        return import_assets_from_project_workbook(db, project)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.post("/projects/{project_id}/sync-template-assets", response_model=asset_schema.AssetImportResult)
def sync_project_template_assets(project_id: int, db: Session = Depends(get_db)):
    return sync_project_workbook_assets(project_id, db)


@router.get("/projects/{project_id}/public")
def read_project_public(project_id: int, db: Session = Depends(get_db)):
    project = db.get(Project, project_id)
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    service_history = (
        db.query(ServiceJob)
        .filter(ServiceJob.project_id == project_id)
        .order_by(ServiceJob.service_date.desc())
        .limit(30)
        .all()
    )
    reports = (
        db.query(Report)
        .join(ServiceJob, Report.service_id == ServiceJob.id)
        .filter(ServiceJob.project_id == project_id)
        .order_by(Report.created_at.desc())
        .all()
    )

    return {
        "customer": {
            "id": project.customer.id,
            "name": project.customer.name,
            "contact_name": project.customer.contact_name,
            "phone": project.customer.phone,
            "email": project.customer.email,
            "address": project.customer.address,
        },
        "project": {
            "id": project.id,
            "name": project.name,
            "location": project.location,
            "description": project.description,
            "template_file_path": project.template_file_path,
            "project_workbook_file_path": project.project_workbook_file_path,
            "created_at": project.created_at,
        },
        "panels": [
            {
                "id": panel.id,
                "panel_code": panel.panel_code,
                "panel_name": panel.panel_name,
                "serial_number": panel.serial_number,
                "location_note": panel.location_note,
                "loops": [
                    {
                        "id": loop.id,
                        "loop_code": loop.loop_code,
                        "loop_name": loop.loop_name,
                        "converter_name": loop.converter_name,
                        "converter_ip": loop.converter_ip,
                        "mac_address": loop.mac_address,
                        "meters": [
                            {
                                "id": meter.id,
                                "meter_code": meter.meter_code,
                                "meter_name": meter.meter_name,
                                "serial_number": meter.serial_number,
                                "device_address": meter.device_address,
                                "model": meter.model,
                                "ct_ratio": meter.ct_ratio,
                                "baud_rate": meter.baud_rate,
                                "status": meter.status,
                            }
                            for meter in loop.meters
                        ],
                    }
                    for loop in panel.loops
                ],
            }
            for panel in project.panels
        ],
        "service_history": [
            {
                "id": job.id,
                "service_type": job.service_type,
                "service_date": job.service_date,
                "status": job.status,
                "note": job.note,
                "engineer_name": job.engineer_name,
                "meter_id": job.meter_id,
                "photos": [photo.file_path for photo in job.photos],
                "reports": [report.file_path for report in job.reports],
            }
            for job in service_history
        ],
        "reports": [
            {
                "id": report.id,
                "service_id": report.service_id,
                "file_path": report.file_path,
                "created_at": report.created_at,
            }
            for report in reports
        ],
    }


@router.get("/templates/")
def list_available_templates():
    """List all available Excel template files in the templates directory."""
    project_uploads_root = UPLOADS_DIR / "projects"
    if not TEMPLATES_DIR.exists() and not project_uploads_root.exists():
        return []
    
    templates = []
    project_template_files = [
        file_path
        for file_path in project_uploads_root.rglob("*.xls*")
        if file_path.is_file() and file_path.parent.name == "templates"
    ]
    for file_path in list(TEMPLATES_DIR.rglob("*.xls*")) + project_template_files:
        if not file_path.is_file():
            continue
        templates.append({
            "name": file_path.name,
            "path": build_public_file_path(UPLOADS_DIR, file_path),
            "size": file_path.stat().st_size,
            "modified": datetime.fromtimestamp(file_path.stat().st_mtime).isoformat()
        })
    return templates


@router.get("/dashboard/stats")
def dashboard_stats(db: Session = Depends(get_db)):
    return {
        "customers": db.query(func.count(Customer.id)).scalar() or 0,
        "projects": db.query(func.count(Project.id)).scalar() or 0,
        "panels": db.query(func.count(Panel.id)).scalar() or 0,
        "meters": db.query(func.count(Meter.id)).scalar() or 0,
        "service_jobs": db.query(func.count(ServiceJob.id)).scalar() or 0,
        "completed_services": db.query(func.count(ServiceJob.id)).filter(ServiceJob.status == "Completed").scalar() or 0,
        "reports": db.query(func.count(Report.id)).scalar() or 0,
    }


# ─── Photo Captions from Template ─────────────────────────────────────────────

@router.get("/projects/{project_id}/photo-captions")
def get_project_photo_captions(project_id: int, db: Session = Depends(get_db)):
    """Get photo caption labels from the project's template file."""
    project = db.get(Project, project_id)
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Resolve template path
    from ..services.report_generator import _resolve_template_path_for_project
    template_path = _resolve_template_path_for_project(project)
    
    if not template_path or not template_path.exists():
        # Return default captions if no template
        return {
            "captions": ["รูปการทำความสะอาดมิเตอร์", "รูปหน้าจอมิเตอร์", "รูปบัตรกำกับมิเตอร์"],
            "source": "default"
        }
    
    captions = extract_template_photo_captions(template_path)
    return {
        "captions": captions,
        "source": "template",
        "template_name": template_path.name
    }


# ─── Project Checklist Templates ─────────────────────────────────────────────

PROJECT_CHECKLIST_STORE = {}  # In-memory store for project checklists (key: project_id)

@router.get("/projects/{project_id}/checklist-template")
def get_project_checklist_template(project_id: int, db: Session = Depends(get_db)):
    """Get custom checklist template for a project. Falls back to default if not set."""
    project = db.get(Project, project_id)
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Check if project has custom checklist
    if project_id in PROJECT_CHECKLIST_STORE:
        return {
            "template": PROJECT_CHECKLIST_STORE[project_id],
            "source": "custom",
            "project_id": project_id
        }
    
    # Try to extract from template file
    from ..services.report_generator import _resolve_template_path_for_project
    template_path = _resolve_template_path_for_project(project)
    
    if template_path and template_path.exists():
        items = extract_template_checklist_items(template_path)
        if items:
            return {
                "template": {
                    "job_type": "PM",
                    "name": f"{project.name} - Checklist",
                    "topics": items
                },
                "source": "template",
                "project_id": project_id
            }
    
    # Return default checklist
    return {
        "template": DEFAULT_CHECKLIST_TEMPLATES.get("PM"),
        "source": "default",
        "project_id": project_id
    }


@router.put("/projects/{project_id}/checklist-template")
def update_project_checklist_template(
    project_id: int,
    template_data: dict,
    db: Session = Depends(get_db)
):
    """Save custom checklist template for a project."""
    project = db.get(Project, project_id)
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Validate template structure
    if "topics" not in template_data:
        raise HTTPException(status_code=400, detail="Template must have 'topics' array")
    
    # Store the custom template
    PROJECT_CHECKLIST_STORE[project_id] = {
        "job_type": template_data.get("job_type", "PM"),
        "name": template_data.get("name", f"{project.name} - Custom"),
        "topics": template_data["topics"]
    }
    
    return {
        "status": "saved",
        "project_id": project_id,
        "template": PROJECT_CHECKLIST_STORE[project_id]
    }


@router.delete("/projects/{project_id}/checklist-template")
def reset_project_checklist_template(project_id: int, db: Session = Depends(get_db)):
    """Reset project checklist to default (remove custom template)."""
    project = db.get(Project, project_id)
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")
    
    if project_id in PROJECT_CHECKLIST_STORE:
        del PROJECT_CHECKLIST_STORE[project_id]
    
    return {
        "status": "reset",
        "project_id": project_id,
        "message": "Checklist template reset to default"
    }


@router.post("/projects/{project_id}/prepare-report-template")
def prepare_report_template(project_id: int, db: Session = Depends(get_db)):
    """Pre-generate Excel template with sheets for all meters, ordered by loop and meter name."""
    import re
    from pathlib import Path
    from openpyxl import load_workbook
    from ..services.excel_import import _resolve_project_workbook_path, get_assets_from_oakwood_workbook
    
    project = db.get(Project, project_id)
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Find template path (resolve /uploads/ prefix)
    template_path = None
    if project.template_file_path:
        relative = project.template_file_path.removeprefix("/uploads/")
        tp = UPLOADS_DIR / relative
        if tp.exists():
            template_path = tp
    if not template_path and DEFAULT_REPORT_TEMPLATE_PATH:
        tp = Path(DEFAULT_REPORT_TEMPLATE_PATH)
        if tp.exists():
            template_path = tp
    if not template_path:
        raise HTTPException(status_code=404, detail="Template not found")
    
    # Find workbook path using existing resolver
    workbook_path = _resolve_project_workbook_path(project)
    if not workbook_path or not workbook_path.exists():
        raise HTTPException(status_code=404, detail="Workbook not found")
    
    # Read meters from workbook
    try:
        from ..services.excel_import import get_assets_from_oakwood_workbook
        data_wb = load_workbook(workbook_path, read_only=True, data_only=True)
        loops = get_assets_from_oakwood_workbook(data_wb)
        data_wb.close()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to read workbook: {str(e)}")
    
    def clean_name(name):
        cleaned = re.sub(r'[^\w\s-]', '', str(name or 'unknown'))
        return cleaned.strip().replace(' ', '_')[:15]
    
    # Collect all meters
    all_meters = []
    for loop in loops:
        for meter in loop.get("meters", []):
            all_meters.append({
                "meter_code": meter.get("meter_code", ""),
                "meter_name": meter.get("meter_name", meter.get("meter_code", "")),
                "loop_name": loop.get("loop_name", "Unknown"),
                "loop_index": loop.get("loop_index", 0),
            })
    
    all_meters.sort(key=lambda m: (m["loop_index"], m["meter_code"]))
    
    # Create output path
    output_dir = project_upload_dir(UPLOADS_DIR, project, "reports")
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "template_prepared.xlsx"
    
    # Load template and copy sheets
    workbook = load_workbook(template_path)
    
    sheet_copies = []
    for idx, meter in enumerate(all_meters):
        loop_clean = clean_name(meter["loop_name"])
        meter_clean = clean_name(meter["meter_name"])
        
        if "Page (4)" in workbook.sheetnames:
            name = f"{idx+1:03d}_{loop_clean}_{meter_clean}_P4"[:31]
            copy = workbook.copy_worksheet(workbook["Page (4)"])
            copy.title = name
            sheet_copies.append(name)
        
        if "Page (5)" in workbook.sheetnames:
            name = f"{idx+1:03d}_{loop_clean}_{meter_clean}_P5"[:31]
            copy = workbook.copy_worksheet(workbook["Page (5)"])
            copy.title = name
            sheet_copies.append(name)
    
    workbook.save(output_path)
    workbook.close()
    
    return {
        "status": "success",
        "project_id": project_id,
        "message": f"Created {len(sheet_copies)} sheets for {len(all_meters)} meters",
        "file_path": build_public_file_path(UPLOADS_DIR, output_path),
        "meters": [{"code": m["meter_code"], "name": m["meter_name"], "loop": m["loop_name"]} for m in all_meters]
    }


@router.post("/projects/{project_id}/meter-drafts")
def save_meter_drafts(project_id: int, payload: dict, db: Session = Depends(get_db)):
    """Save meter report draft data to JSON + update xlsx sheets."""
    import json, logging, re
    from openpyxl import load_workbook
    logger = logging.getLogger(__name__)
    
    project = db.get(Project, project_id)
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")
    
    drafts_dir = project_upload_dir(UPLOADS_DIR, project, "documents", "report_drafts")
    draft_path = drafts_dir / "meter_drafts.json"
    
    # 1. Save to JSON (fast, reliable)
    existing = {}
    if draft_path.exists():
        try:
            existing = json.loads(draft_path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            existing = {}
    
    meter_data = payload.get("meter_data", {})
    existing.update(meter_data)
    
    draft_path.write_text(
        json.dumps(existing, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    
    logger.info(f"[DRAFT] JSON saved for project {project_id}: {len(meter_data)} meters updated, {len(existing)} total")
    
    # 2. Also write data into template_prepared.xlsx sheets
    xlsx_updated = 0
    reports_dir = project_upload_dir(UPLOADS_DIR, project, "reports")
    xlsx_path = reports_dir / "template_prepared.xlsx"
    if xlsx_path.exists():
        try:
            wb = load_workbook(xlsx_path)
            for meter_code, data in meter_data.items():
                if not data or not isinstance(data, dict):
                    continue
                # Find matching sheet by meter_code in sheet name
                clean_code = re.sub(r'[^\w]', '', meter_code).lower()
                for sheet_name in wb.sheetnames:
                    clean_sheet = re.sub(r'[^\w]', '', sheet_name).lower()
                    if clean_code in clean_sheet and 'p4' in clean_sheet:
                        ws = wb[sheet_name]
                        # Write energy reading
                        if data.get('energy_reading'):
                            ws['F10'] = str(data['energy_reading'])
                        # Write online/offline status
                        ws['F11'] = 'Online' if data.get('online_status', True) else 'Offline'
                        # Write accuracy
                        ws['F12'] = data.get('accuracy_status', 'Pass')
                        # Write comment
                        if data.get('comment'):
                            ws['F13'] = str(data['comment'])
                        # Write checklist items
                        checklist = data.get('checklist', {})
                        row = 16  # Start row for checklist
                        for topic_id, item in checklist.items():
                            if isinstance(item, dict):
                                ws[f'B{row}'] = item.get('label', topic_id)
                                ws[f'E{row}'] = item.get('status', 'N/A')
                                ws[f'F{row}'] = item.get('note', '')
                                row += 1
                        xlsx_updated += 1
                        break
            wb.save(xlsx_path)
            wb.close()
            logger.info(f"[DRAFT] XLSX updated: {xlsx_updated} meter sheets written")
        except Exception as e:
            logger.warning(f"[DRAFT] XLSX update failed (non-critical): {e}")
    
    return {
        "status": "saved",
        "project_id": project_id,
        "meters_saved": len(meter_data),
        "total_meters": len(existing),
        "xlsx_updated": xlsx_updated,
    }


@router.get("/projects/{project_id}/meter-drafts")
def load_meter_drafts(project_id: int, db: Session = Depends(get_db)):
    """Load saved meter report draft data."""
    import json, logging
    logger = logging.getLogger(__name__)
    
    project = db.get(Project, project_id)
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")
    
    drafts_dir = project_upload_dir(UPLOADS_DIR, project, "documents", "report_drafts")
    draft_path = drafts_dir / "meter_drafts.json"
    
    if not draft_path.exists():
        logger.info(f"No meter drafts found for project {project_id}")
        return {"project_id": project_id, "meter_data": {}}
    
    try:
        data = json.loads(draft_path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        data = {}
    
    logger.info(f"Loaded meter drafts for project {project_id}: {len(data)} meters")
    return {"project_id": project_id, "meter_data": data}


# Import default templates for fallback
from ..services.report_generator import DEFAULT_CHECKLIST_TEMPLATES

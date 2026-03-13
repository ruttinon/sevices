from io import BytesIO

import qrcode
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from ..config import DOCUMENTS_DIR
from ..database import get_db
from ..models import Loop, Meter, Panel, Project, Report, ServiceJob
from ..schemas import asset_schema
from ..services.asset_export import export_project_assets_to_excel
from ..services.excel_import import (
    get_assets_from_oakwood_workbook,
    import_assets_from_excel,
    import_assets_from_project_workbook,
)
from ..services.file_upload import save_upload_file

from ..services.qr_scan import build_qr_payload
from ..utils.project_files import project_upload_dir
from ..utils.validators import validate_meter_status

router = APIRouter(tags=["assets"])


def _stream_qr_image(payload: str, filename: str):
    image = qrcode.make(payload)
    buffer = BytesIO()
    image.save(buffer, format="PNG")
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="image/png",
        headers={"Content-Disposition": f'inline; filename="{filename}"'},
    )


@router.get("/panels", response_model=list[asset_schema.PanelResponse])
def read_panels(project_id: int | None = None, db: Session = Depends(get_db)):
    query = db.query(Panel)
    if project_id is not None:
        query = query.filter(Panel.project_id == project_id)
    return query.order_by(Panel.panel_code.asc()).all()


@router.post("/panels", response_model=asset_schema.PanelResponse)
def create_panel(panel: asset_schema.PanelCreate, db: Session = Depends(get_db)):
    if db.get(Project, panel.project_id) is None:
        raise HTTPException(status_code=404, detail="Project not found")

    db_panel = Panel(**panel.model_dump())
    db.add(db_panel)
    db.commit()
    db.refresh(db_panel)
    return db_panel


@router.put("/panels/{panel_id}", response_model=asset_schema.PanelResponse)
def update_panel(panel_id: int, panel_update: asset_schema.PanelUpdate, db: Session = Depends(get_db)):
    panel = db.get(Panel, panel_id)
    if panel is None:
        raise HTTPException(status_code=404, detail="Panel not found")

    update_data = panel_update.model_dump(exclude_unset=True)
    if "project_id" in update_data and db.get(Project, update_data["project_id"]) is None:
        raise HTTPException(status_code=404, detail="Project not found")

    for field, value in update_data.items():
        setattr(panel, field, value)

    db.commit()
    db.refresh(panel)
    return panel


@router.delete("/panels/{panel_id}")
def delete_panel(panel_id: int, db: Session = Depends(get_db)):
    panel = db.get(Panel, panel_id)
    if panel is None:
        raise HTTPException(status_code=404, detail="Panel not found")
    if any(meter.service_jobs for loop in panel.loops for meter in loop.meters):
        raise HTTPException(status_code=400, detail="This panel has service history and cannot be deleted")

    db.delete(panel)
    db.commit()
    return {"status": "deleted"}


@router.get("/panels/{panel_id}", response_model=asset_schema.PanelResponse)
def read_panel(panel_id: int, db: Session = Depends(get_db)):
    panel = db.get(Panel, panel_id)
    if panel is None:
        raise HTTPException(status_code=404, detail="Panel not found")
    return panel


@router.get("/panels/{panel_id}/qr")
def generate_panel_qr(panel_id: int, db: Session = Depends(get_db)):
    panel = db.get(Panel, panel_id)
    if panel is None:
        raise HTTPException(status_code=404, detail="Panel not found")

    payload = build_qr_payload("panel", panel.id, panel.project_id)
    return _stream_qr_image(payload, f"panel-{panel.panel_code}.png")


@router.get("/projects/{project_id}/qr")
def generate_project_qr(project_id: int, db: Session = Depends(get_db)):
    project = db.get(Project, project_id)
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    payload = build_qr_payload("project", project.id, project.id)
    return _stream_qr_image(payload, f"project-{project.id}.png")


@router.get("/loops/{loop_id}/qr")
def generate_loop_qr(loop_id: int, db: Session = Depends(get_db)):
    loop = db.get(Loop, loop_id)
    if loop is None:
        raise HTTPException(status_code=404, detail="Loop not found")

    payload = build_qr_payload("loop", loop.id, loop.panel.project_id)
    return _stream_qr_image(payload, f"loop-{loop.loop_code}.png")


@router.get("/meters/{meter_id}/qr")
def generate_meter_qr(meter_id: int, db: Session = Depends(get_db)):
    meter = db.get(Meter, meter_id)
    if meter is None:
        raise HTTPException(status_code=404, detail="Meter not found")

    payload = build_qr_payload("meter", meter.id, meter.loop.panel.project_id)
    return _stream_qr_image(payload, f"meter-{meter.meter_code}.png")


@router.get("/panels/{panel_id}/public")
def read_panel_public(panel_id: int, db: Session = Depends(get_db)):
    panel = db.get(Panel, panel_id)
    if panel is None:
        raise HTTPException(status_code=404, detail="Panel not found")

    service_history = (
        db.query(ServiceJob)
        .join(Meter, ServiceJob.meter_id == Meter.id)
        .join(Loop, Meter.loop_id == Loop.id)
        .filter(Loop.panel_id == panel_id)
        .order_by(ServiceJob.service_date.desc())
        .all()
    )
    reports = (
        db.query(Report)
        .join(ServiceJob, Report.service_id == ServiceJob.id)
        .join(Meter, ServiceJob.meter_id == Meter.id)
        .join(Loop, Meter.loop_id == Loop.id)
        .filter(Loop.panel_id == panel_id)
        .order_by(Report.created_at.desc())
        .all()
    )

    return {
        "customer": {
            "id": panel.project.customer.id,
            "name": panel.project.customer.name,
            "contact_name": panel.project.customer.contact_name,
            "phone": panel.project.customer.phone,
            "email": panel.project.customer.email,
            "address": panel.project.customer.address,
        },
        "project": {
            "id": panel.project.id,
            "name": panel.project.name,
            "location": panel.project.location,
            "description": panel.project.description,
            "template_file_path": panel.project.template_file_path,
            "project_workbook_file_path": panel.project.project_workbook_file_path,
        },
        "panel": {
            "id": panel.id,
            "panel_code": panel.panel_code,
            "panel_name": panel.panel_name,
            "serial_number": panel.serial_number,
            "location_note": panel.location_note,
        },
        "loops": [
            {
                "id": loop.id,
                "loop_code": loop.loop_code,
                "loop_name": loop.loop_name,
                "converter_name": loop.converter_name,
                "converter_ip": loop.converter_ip,
                "mac_address": loop.mac_address,
                "meters": [
                    {
                        "id": meter.id,
                        "meter_code": meter.meter_code,
                        "meter_name": meter.meter_name,
                        "serial_number": meter.serial_number,
                        "device_address": meter.device_address,
                        "model": meter.model,
                        "ct_ratio": meter.ct_ratio,
                        "baud_rate": meter.baud_rate,
                        "status": meter.status,
                    }
                    for meter in loop.meters
                ],
            }
            for loop in panel.loops
        ],
        "service_history": [
            {
                "id": job.id,
                "meter_id": job.meter_id,
                "meter_name": job.meter.meter_name if job.meter else None,
                "service_type": job.service_type,
                "service_date": job.service_date,
                "status": job.status,
                "note": job.note,
                "engineer_name": job.engineer_name,
                "reports": [report.file_path for report in job.reports],
            }
            for job in service_history
        ],
        "reports": [
            {
                "id": report.id,
                "service_id": report.service_id,
                "file_path": report.file_path,
                "created_at": report.created_at,
            }
            for report in reports
        ],
    }


@router.get("/loops", response_model=list[asset_schema.LoopResponse])
def read_loops(panel_id: int | None = None, db: Session = Depends(get_db)):
    query = db.query(Loop)
    if panel_id is not None:
        query = query.filter(Loop.panel_id == panel_id)
    return query.order_by(Loop.loop_code.asc()).all()


@router.post("/loops", response_model=asset_schema.LoopResponse)
def create_loop(loop: asset_schema.LoopCreate, db: Session = Depends(get_db)):
    if db.get(Panel, loop.panel_id) is None:
        raise HTTPException(status_code=404, detail="Panel not found")

    db_loop = Loop(**loop.model_dump())
    db.add(db_loop)
    db.commit()
    db.refresh(db_loop)
    return db_loop


@router.put("/loops/{loop_id}", response_model=asset_schema.LoopResponse)
def update_loop(loop_id: int, loop_update: asset_schema.LoopUpdate, db: Session = Depends(get_db)):
    loop = db.get(Loop, loop_id)
    if loop is None:
        raise HTTPException(status_code=404, detail="Loop not found")

    update_data = loop_update.model_dump(exclude_unset=True)
    if "panel_id" in update_data and db.get(Panel, update_data["panel_id"]) is None:
        raise HTTPException(status_code=404, detail="Panel not found")

    for field, value in update_data.items():
        setattr(loop, field, value)

    db.commit()
    db.refresh(loop)
    return loop


@router.delete("/loops/{loop_id}")
def delete_loop(loop_id: int, db: Session = Depends(get_db)):
    loop = db.get(Loop, loop_id)
    if loop is None:
        raise HTTPException(status_code=404, detail="Loop not found")
    if any(meter.service_jobs for meter in loop.meters):
        raise HTTPException(status_code=400, detail="This loop has service history and cannot be deleted")

    db.delete(loop)
    db.commit()
    return {"status": "deleted"}


@router.get("/meters", response_model=list[asset_schema.MeterResponse])
def read_meters(
    project_id: int | None = None,
    panel_id: int | None = None,
    loop_id: int | None = None,
    db: Session = Depends(get_db),
):
    query = db.query(Meter)
    if loop_id is not None:
        query = query.filter(Meter.loop_id == loop_id)
    elif panel_id is not None:
        query = query.join(Loop).filter(Loop.panel_id == panel_id)
    elif project_id is not None:
        query = query.join(Loop).join(Panel).filter(Panel.project_id == project_id)
    return query.order_by(Meter.meter_code.asc()).all()


@router.post("/meters", response_model=asset_schema.MeterResponse)
def create_meter(meter: asset_schema.MeterCreate, db: Session = Depends(get_db)):
    if db.get(Loop, meter.loop_id) is None:
        raise HTTPException(status_code=404, detail="Loop not found")

    try:
        validate_meter_status(meter.status)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    db_meter = Meter(**meter.model_dump())
    db.add(db_meter)
    db.commit()
    db.refresh(db_meter)
    return db_meter


@router.get("/projects/{project_id}/workbook-assets")
def get_project_workbook_assets(project_id: int, db: Session = Depends(get_db)):
    from openpyxl import load_workbook
    from ..services.excel_import import _resolve_project_workbook_path

    project = db.get(Project, project_id)
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    workbook_path = _resolve_project_workbook_path(project)
    print(f"DEBUG: Opening workbook for project {project_id} at {workbook_path}")
    
    if not workbook_path or not workbook_path.exists():
        # Return empty list instead of 404 to prevent frontend crashes
        return []

    try:
        workbook = load_workbook(workbook_path, data_only=True)
        print(f"DEBUG: Workbook loaded. Sheets: {workbook.sheetnames}")
        return get_assets_from_oakwood_workbook(workbook)
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to read or parse workbook: {e}")


@router.put("/meters/{meter_id}", response_model=asset_schema.MeterResponse)
def update_meter(meter_id: int, meter_update: asset_schema.MeterUpdate, db: Session = Depends(get_db)):
    meter = db.get(Meter, meter_id)
    if meter is None:
        raise HTTPException(status_code=404, detail="Meter not found")

    update_data = meter_update.model_dump(exclude_unset=True)
    if "loop_id" in update_data and db.get(Loop, update_data["loop_id"]) is None:
        raise HTTPException(status_code=404, detail="Loop not found")
    if "status" in update_data:
        try:
            validate_meter_status(update_data["status"])
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    for field, value in update_data.items():
        setattr(meter, field, value)

    db.commit()
    db.refresh(meter)
    return meter


@router.delete("/meters/{meter_id}")
def delete_meter(meter_id: int, db: Session = Depends(get_db)):
    meter = db.get(Meter, meter_id)
    if meter is None:
        raise HTTPException(status_code=404, detail="Meter not found")
    if meter.service_jobs:
        raise HTTPException(status_code=400, detail="This meter has service history and cannot be deleted")

    db.delete(meter)
    db.commit()
    return {"status": "deleted"}


@router.post("/projects/{project_id}/photos/upload")
def upload_project_photo(
    project_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    project = db.get(Project, project_id)
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    from ..config import UPLOADS_DIR
    from ..services.file_upload import save_upload
    from ..utils.project_files import project_upload_dir

    photos_dir = project_upload_dir(UPLOADS_DIR, project, "photos")
    public_path = save_upload(file, photos_dir, prefix="PHOTO", keep_original_name=True)

    return {"file_path": public_path}


@router.get("/meters/{meter_id}")
def read_meter_detail(meter_id: int, db: Session = Depends(get_db)):
    meter = db.get(Meter, meter_id)
    if meter is None:
        raise HTTPException(status_code=404, detail="Meter not found")

    loop = meter.loop
    panel = loop.panel
    project = panel.project

    return {
        "meter": {
            "id": meter.id,
            "meter_code": meter.meter_code,
            "meter_name": meter.meter_name,
            "serial_number": meter.serial_number,
            "device_address": meter.device_address,
            "model": meter.model,
            "ct_ratio": meter.ct_ratio,
            "baud_rate": meter.baud_rate,
            "status": meter.status,
        },
        "loop": {
            "id": loop.id,
            "loop_code": loop.loop_code,
            "loop_name": loop.loop_name,
            "converter_name": loop.converter_name,
            "converter_ip": loop.converter_ip,
            "mac_address": loop.mac_address,
        },
        "panel": {
            "id": panel.id,
            "panel_code": panel.panel_code,
            "panel_name": panel.panel_name,
            "serial_number": panel.serial_number,
            "location_note": panel.location_note,
        },
        "project": {
            "id": project.id,
            "name": project.name,
            "location": project.location,
            "customer_id": project.customer_id,
        },
        "service_history": [
            {
                "id": service.id,
                "service_type": service.service_type,
                "service_date": service.service_date,
                "status": service.status,
                "engineer_name": service.engineer_name,
                "note": service.note,
                "photos": [photo.file_path for photo in service.photos],
                "reports": [report.file_path for report in service.reports],
            }
            for service in sorted(meter.service_jobs, key=lambda item: item.service_date, reverse=True)
        ],
    }


@router.post("/assets/import", response_model=asset_schema.AssetImportResult)
def import_asset_excel(
    project_id: int = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    project = db.get(Project, project_id)
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    import_dir = project_upload_dir(DOCUMENTS_DIR.parent, project, "documents", "imports")
    workbook_path = save_upload_file(file, import_dir, prefix="IMPORT", keep_original_name=True)

    try:
        return import_assets_from_excel(db, workbook_path, project_id)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    finally:
        workbook_path.unlink(missing_ok=True)


@router.post("/projects/{project_id}/assets/sync-template", response_model=asset_schema.AssetImportResult)
def sync_assets_from_project_template(project_id: int, db: Session = Depends(get_db)):
    project = db.get(Project, project_id)
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    try:
        return import_assets_from_project_workbook(db, project)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.post("/projects/{project_id}/assets/export-excel", response_model=asset_schema.AssetExportResult)
def export_assets_to_excel(project_id: int, db: Session = Depends(get_db)):
    project = db.get(Project, project_id)
    if project is None:
        raise HTTPException(status_code=404, detail="Project not found")

    try:
        return export_project_assets_to_excel(db, project)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

from pathlib import Path
import shutil

from sqlalchemy import create_engine, inspect
from sqlalchemy.orm import declarative_base
from sqlalchemy.orm import sessionmaker
from .config import DATABASE_URL, UPLOADS_DIR

engine = create_engine(
    DATABASE_URL, connect_args={"check_same_thread": False}
)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)

Base = declarative_base()


REQUIRED_SCHEMA = {
    "customers": {"id", "name", "contact_name", "phone", "email", "address"},
    "projects": {"id", "customer_id", "name", "location", "description", "template_file_path", "created_at"},
    "panels": {"id", "project_id", "panel_code", "panel_name", "serial_number", "location_note"},
    "loops": {"id", "panel_id", "loop_code", "loop_name", "converter_name", "converter_ip", "mac_address"},
    "meters": {"id", "loop_id", "meter_code", "meter_name", "serial_number", "device_address", "model", "ct_ratio", "baud_rate", "status"},
    "service_jobs": {"id", "project_id", "meter_id", "engineer_id", "engineer_name", "service_type", "service_date", "status", "note", "checklist_items", "report_draft", "created_at"},
    "service_photos": {"id", "service_id", "file_path", "caption", "created_at"},
    "reports": {"id", "service_id", "file_path", "created_at"},
}


def _requires_reset() -> bool:
    inspector = inspect(engine)
    existing_tables = set(inspector.get_table_names())

    for table_name, required_columns in REQUIRED_SCHEMA.items():
        if table_name not in existing_tables:
            continue
        existing_columns = {column["name"] for column in inspector.get_columns(table_name)}
        if not required_columns.issubset(existing_columns):
            return True

    return False


def _ensure_optional_project_columns() -> None:
    inspector = inspect(engine)
    existing_tables = set(inspector.get_table_names())
    
    if "projects" in existing_tables:
        existing_columns = {column["name"] for column in inspector.get_columns("projects")}
        if "project_workbook_file_path" not in existing_columns:
            with engine.begin() as connection:
                connection.exec_driver_sql("ALTER TABLE projects ADD COLUMN project_workbook_file_path VARCHAR")

    if "service_jobs" in existing_tables:
        existing_columns = {column["name"] for column in inspector.get_columns("service_jobs")}
        if "report_draft" not in existing_columns:
            with engine.begin() as connection:
                connection.exec_driver_sql("ALTER TABLE service_jobs ADD COLUMN report_draft JSON")

    if "reports" in existing_tables:
        existing_columns = {column["name"] for column in inspector.get_columns("reports")}
        if "report_date" not in existing_columns:
            with engine.begin() as connection:
                connection.exec_driver_sql("ALTER TABLE reports ADD COLUMN report_date DATETIME")


def _absolute_path_from_public(public_path: str) -> Path:
    return UPLOADS_DIR / public_path.removeprefix("/uploads/")


def _looks_like_oakwood_workbook(path: Path) -> bool:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True)
    except Exception:
        return "maintenance agreement" in path.name.lower() or "oakwood" in path.name.lower()

    try:
        sheet_names = set(workbook.sheetnames)
        return "Cover" in sheet_names and any(name.lower().startswith("loop") for name in workbook.sheetnames)
    finally:
        workbook.close()


def _migrate_project_file_roles() -> None:
    from .models import Project

    db = SessionLocal()
    try:
        projects = db.query(Project).all()
        changed = False
        for project in projects:
            if project.project_workbook_file_path:
                continue
            if not project.template_file_path:
                continue

            candidate_path = _absolute_path_from_public(project.template_file_path)
            if not candidate_path.exists():
                continue

            if _looks_like_oakwood_workbook(candidate_path):
                project.project_workbook_file_path = project.template_file_path
                project.template_file_path = None
                changed = True

        if changed:
            db.commit()
    finally:
        db.close()


def _relocate_project_uploads() -> None:
    import re

    from .config import DOCUMENTS_DIR, PHOTOS_DIR, REPORT_DRAFTS_DIR, REPORTS_DIR, SYSTEM_DOCUMENTS_DIR, SYSTEM_UPLOADS_DIR
    from .models import Project, Report, ServiceJob, ServicePhoto
    from .utils.helpers import build_public_file_path, generate_qr_code
    from .utils.project_files import project_upload_dir, slugify_path_fragment

    def move_into_project_dir(source_path: Path, target_dir: Path) -> Path:
        target_dir.mkdir(parents=True, exist_ok=True)
        destination = target_dir / source_path.name
        if destination.exists() and destination.resolve() != source_path.resolve():
            destination = target_dir / f"{source_path.stem}-{generate_qr_code('FILE')[-6:].lower()}{source_path.suffix}"
        if source_path.resolve() != destination.resolve():
            try:
                shutil.move(str(source_path), destination)
            except PermissionError:
                shutil.copy2(source_path, destination)
        return destination

    def infer_project_from_name(file_name: str, projects: list) -> object | None:
        match = re.search(r"service-(\d+)", file_name.lower())
        if match:
            service_job = db.get(ServiceJob, int(match.group(1)))
            if service_job is not None:
                return service_job.project

        match = re.search(r"project-(\d{4,})", file_name.lower())
        if match:
            project = db.get(Project, int(match.group(1)))
            if project is not None:
                return project

        lowered_name = file_name.lower()
        for project in projects:
            slug = slugify_path_fragment(project.name, "project")
            if slug and slug.lower() in lowered_name:
                return project
        return None

    def move_public_path(public_path: str, target_dir: Path) -> str:
        source_path = _absolute_path_from_public(public_path)
        if not source_path.exists():
            return public_path
        destination = move_into_project_dir(source_path, target_dir)
        return build_public_file_path(UPLOADS_DIR, destination)

    db = SessionLocal()
    try:
        changed = False
        projects = db.query(Project).all()
        for project in projects:
            project_files = [
                ("template_file_path", "templates"),
                ("project_workbook_file_path", "workbooks"),
            ]
            for attribute_name, category in project_files:
                public_path = getattr(project, attribute_name)
                if not public_path:
                    continue
                target_dir = project_upload_dir(UPLOADS_DIR, project, "documents", category)
                new_public_path = move_public_path(public_path, target_dir)
                if public_path != new_public_path:
                    setattr(project, attribute_name, new_public_path)
                    changed = True

        service_jobs = db.query(ServiceJob).all()
        for service_job in service_jobs:
            legacy_draft_path = REPORT_DRAFTS_DIR / f"service-{service_job.id:06d}.json"
            if not legacy_draft_path.exists():
                legacy_candidates = list((DOCUMENTS_DIR / "projects").glob(f"*/report_drafts/{legacy_draft_path.name}"))
                legacy_candidates.extend((UPLOADS_DIR / "projects").glob(f"project-*/documents/report_drafts/{legacy_draft_path.name}"))
                if not legacy_candidates:
                    continue
                legacy_draft_path = legacy_candidates[0]
            target_dir = project_upload_dir(UPLOADS_DIR, service_job.project, "documents", "report_drafts")
            move_into_project_dir(legacy_draft_path, target_dir)

        for photo in db.query(ServicePhoto).all():
            service_job = db.get(ServiceJob, photo.service_id)
            if service_job is None or not photo.file_path:
                continue
            target_dir = project_upload_dir(UPLOADS_DIR, service_job.project, "photos")
            new_public_path = move_public_path(photo.file_path, target_dir)
            if new_public_path != photo.file_path:
                photo.file_path = new_public_path
                changed = True

        for report in db.query(Report).all():
            service_job = db.get(ServiceJob, report.service_id)
            if service_job is None or not report.file_path:
                continue
            target_dir = project_upload_dir(UPLOADS_DIR, service_job.project, "reports")
            new_public_path = move_public_path(report.file_path, target_dir)
            if new_public_path != report.file_path:
                report.file_path = new_public_path
                changed = True

        if REPORTS_DIR.exists():
            for report_path in REPORTS_DIR.rglob("*"):
                if not report_path.is_file():
                    continue
                project = infer_project_from_name(report_path.name, projects)
                if project is not None:
                    target_dir = project_upload_dir(UPLOADS_DIR, project, "reports")
                else:
                    target_dir = SYSTEM_UPLOADS_DIR / "legacy_misc" / "reports"
                move_into_project_dir(report_path, target_dir)

        if PHOTOS_DIR.exists():
            for photo_path in PHOTOS_DIR.rglob("*"):
                if not photo_path.is_file():
                    continue
                project = infer_project_from_name(photo_path.name, projects)
                if project is not None:
                    target_dir = project_upload_dir(UPLOADS_DIR, project, "photos")
                else:
                    target_dir = SYSTEM_UPLOADS_DIR / "legacy_misc" / "photos"
                move_into_project_dir(photo_path, target_dir)

        legacy_project_root = UPLOADS_DIR / "projects"
        for legacy_project_dir in legacy_project_root.glob("project-*"):
            if not legacy_project_dir.is_dir():
                continue
            match = re.match(r"project-(\d{4,})", legacy_project_dir.name.lower())
            project = db.get(Project, int(match.group(1))) if match else None
            if project is None:
                continue
            current_root = project_upload_dir(UPLOADS_DIR, project)
            for item in legacy_project_dir.rglob("*"):
                if not item.is_file():
                    continue
                relative = item.relative_to(legacy_project_dir)
                target_dir = current_root / relative.parent
                move_into_project_dir(item, target_dir)
            try:
                legacy_project_dir.rmdir()
            except OSError:
                pass

        if DOCUMENTS_DIR.exists():
            legacy_asset_exports_dir = DOCUMENTS_DIR / "asset_exports"
            for project_folder in legacy_asset_exports_dir.glob("project-*"):
                if not project_folder.is_dir():
                    continue
                match = re.match(r"project-(\d{4,})", project_folder.name.lower())
                project = db.get(Project, int(match.group(1))) if match else None
                if project is not None:
                    target_dir = project_upload_dir(UPLOADS_DIR, project, "documents", "asset_exports")
                else:
                    target_dir = SYSTEM_DOCUMENTS_DIR / "legacy_misc" / "asset_exports" / project_folder.name
                for item in project_folder.rglob("*"):
                    if item.is_file():
                        move_into_project_dir(item, target_dir)

            for item in DOCUMENTS_DIR.iterdir():
                if item.name in {"projects", "_system", "asset_exports", "report_drafts", "templates", "workbooks"}:
                    continue
                if item.is_dir():
                    continue
                move_into_project_dir(item, SYSTEM_DOCUMENTS_DIR / "legacy_misc")

            legacy_system_dir = DOCUMENTS_DIR / "_system"
            if legacy_system_dir.exists():
                for item in legacy_system_dir.rglob("*"):
                    if not item.is_file():
                        continue
                    relative = item.relative_to(legacy_system_dir)
                    target_dir = SYSTEM_DOCUMENTS_DIR / relative.parent
                    move_into_project_dir(item, target_dir)

        cleanup_roots = [
            REPORTS_DIR,
            PHOTOS_DIR,
            DOCUMENTS_DIR / "projects",
            DOCUMENTS_DIR / "asset_exports",
            DOCUMENTS_DIR / "report_drafts",
            DOCUMENTS_DIR / "templates",
            DOCUMENTS_DIR / "workbooks",
            DOCUMENTS_DIR / "_system",
        ]
        for root_dir in cleanup_roots:
            if not root_dir.exists():
                continue
            for candidate in sorted(root_dir.rglob("*"), reverse=True):
                if candidate.is_dir():
                    try:
                        candidate.rmdir()
                    except OSError:
                        pass
            try:
                root_dir.rmdir()
            except OSError:
                pass

        if changed:
            db.commit()
    finally:
        db.close()


def _seed_demo_data():
    from .config import DEFAULT_PROJECT_WORKBOOK_PATH, DEFAULT_REPORT_TEMPLATE_PATH, UPLOADS_DIR
    from .models import Customer, Loop, Meter, Panel, Project, ServiceJob
    from .utils.helpers import build_public_file_path
    from .utils.project_files import project_upload_dir

    db = SessionLocal()
    try:
        customer = db.query(Customer).filter(Customer.name == "Oakwood Suites Bangkok").first()
        if customer is None:
            customer = Customer(
                name="Oakwood Suites Bangkok",
                contact_name="Property Engineering Team",
                phone="02-210-1234",
                email="engineering@oakwood.local",
                address="Bangkok, Thailand",
            )
            db.add(customer)
            db.flush()

        project = db.query(Project).filter(Project.name == "Oakwood RMS Maintenance").first()
        if project is None:
            project = Project(
                customer_id=customer.id,
                name="Oakwood RMS Maintenance",
                location="Oakwood Suites Bangkok",
                description="Sample project for cabinet scan, customer lookup, and Excel report generation.",
            )
            db.add(project)
            db.flush()

        if not project.template_file_path and DEFAULT_REPORT_TEMPLATE_PATH.exists():
            copied_template_path = project_upload_dir(UPLOADS_DIR, project, "documents", "templates") / DEFAULT_REPORT_TEMPLATE_PATH.name
            if not copied_template_path.exists():
                shutil.copy2(DEFAULT_REPORT_TEMPLATE_PATH, copied_template_path)
            project.template_file_path = build_public_file_path(UPLOADS_DIR, copied_template_path)
            db.flush()

        if not project.project_workbook_file_path and DEFAULT_PROJECT_WORKBOOK_PATH.exists():
            copied_workbook_path = project_upload_dir(UPLOADS_DIR, project, "documents", "workbooks") / DEFAULT_PROJECT_WORKBOOK_PATH.name
            if not copied_workbook_path.exists():
                shutil.copy2(DEFAULT_PROJECT_WORKBOOK_PATH, copied_workbook_path)
            project.project_workbook_file_path = build_public_file_path(UPLOADS_DIR, copied_workbook_path)
            db.flush()

        panel = db.query(Panel).filter(Panel.project_id == project.id, Panel.panel_code == "9DP").first()
        if panel is None:
            panel = Panel(
                project_id=project.id,
                panel_code="9DP",
                panel_name="ESI",
                serial_number="PANEL-9DP-001",
                location_note="Retail Management System (RMS) / 192.168.0.1",
            )
            db.add(panel)
            db.flush()

        loop1 = db.query(Loop).filter(Loop.panel_id == panel.id, Loop.loop_code == "Loop1").first()
        if loop1 is None:
            loop1 = Loop(
                panel_id=panel.id,
                loop_code="Loop1",
                loop_name="Loop1",
                converter_name="AVERA",
                converter_ip="192.168.0.1",
                mac_address="00:26:45:00:FD:95",
            )
            db.add(loop1)
            db.flush()

        loop2 = db.query(Loop).filter(Loop.panel_id == panel.id, Loop.loop_code == "Loop2").first()
        if loop2 is None:
            loop2 = Loop(
                panel_id=panel.id,
                loop_code="Loop2",
                loop_name="Loop2",
                converter_name="AVERA",
                converter_ip="192.168.0.2",
                mac_address="00:26:45:00:FD:96",
            )
            db.add(loop2)
            db.flush()

        sample_meters = [
            ("901", "1177340118", "1"),
            ("902", "1177370235", "2"),
            ("903", "1177340117", "3"),
            ("904", "1177340121", "4"),
            ("905", "1177340120", "5"),
            ("906", "1177340126", "6"),
            ("907", "1177340124", "7"),
            ("908", "1177340125", "8"),
            ("909", "1177340123", "9"),
            ("910", "1177340119", "10"),
            ("911", "1177340113", "11"),
            ("912", "1177340128", "12"),
        ]
        created_meter = None
        for index, (meter_code, serial_number, device_address) in enumerate(sample_meters):
            target_loop = loop1 if index < 6 else loop2
            meter = (
                db.query(Meter)
                .filter(Meter.loop_id == target_loop.id, Meter.meter_code == meter_code)
                .first()
            )
            if meter is None:
                meter = Meter(
                    loop_id=target_loop.id,
                    meter_code=meter_code,
                    meter_name=meter_code,
                    serial_number=serial_number,
                    device_address=device_address,
                    model="CVM-C10",
                    ct_ratio="50/5",
                    baud_rate="9600",
                    status="Active",
                )
                db.add(meter)
                db.flush()
            if meter_code == "901":
                created_meter = meter

        if created_meter is not None:
            seeded_jobs = (
                db.query(ServiceJob)
                .filter(
                    ServiceJob.project_id == project.id,
                    ServiceJob.engineer_name == "Tech Demo",
                    ServiceJob.note == "Seeded maintenance report for Loop1 cabinet.",
                )
                .all()
            )
            for seeded_job in seeded_jobs:
                db.delete(seeded_job)

        db.commit()
    finally:
        db.close()


def init_database():
    from . import models  # noqa: F401

    if _requires_reset():
        Base.metadata.drop_all(bind=engine)
        with engine.begin() as connection:
            connection.exec_driver_sql("DROP TABLE IF EXISTS assets")

    Base.metadata.create_all(bind=engine)
    _ensure_optional_project_columns()
    _migrate_project_file_roles()
    _relocate_project_uploads()
    _seed_demo_data()

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

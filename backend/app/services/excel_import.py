from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from ..config import DEFAULT_PROJECT_WORKBOOK_PATH, UPLOADS_DIR
from ..models import Loop, Meter, Panel, Project


def import_assets_from_excel(db: Session, workbook_path: Path, project_id: int):
    workbook = load_workbook(workbook_path, data_only=False)
    if _is_oakwood_loop_workbook(workbook):
        return _import_oakwood_loop_workbook(db, workbook, project_id)
    return _import_flat_asset_sheet(db, workbook_path, project_id)


def import_assets_from_project_workbook(db: Session, project: Project):
    workbook_path = _resolve_project_workbook_path(project)
    if workbook_path is None or not workbook_path.exists():
        raise ValueError("This project does not have a linked Excel workbook to sync from.")
    return import_assets_from_excel(db, workbook_path, project.id)


def import_assets_from_project_template(db: Session, project: Project):
    return import_assets_from_project_workbook(db, project)


def _import_flat_asset_sheet(db: Session, workbook_path: Path, project_id: int):
    dataframe = pd.read_excel(workbook_path)

    required_columns = {
        "panel_code",
        "panel_name",
        "panel_serial_number",
        "panel_location_note",
        "loop_code",
        "loop_name",
        "converter_name",
        "converter_ip",
        "mac_address",
        "meter_code",
        "meter_name",
        "meter_serial_number",
        "device_address",
        "model",
        "ct_ratio",
        "baud_rate",
        "status",
    }
    missing_columns = required_columns - set(dataframe.columns)
    if missing_columns:
        raise ValueError(f"Missing required columns: {', '.join(sorted(missing_columns))}")

    summary = _empty_import_summary()
    panel_cache: dict[str, Panel] = {}
    loop_cache: dict[tuple[int, str], Loop] = {}

    for _, row in dataframe.iterrows():
        panel_code = _string_value(row["panel_code"])
        loop_code = _string_value(row["loop_code"])
        meter_code = _string_value(row["meter_code"])

        if not panel_code or not loop_code or not meter_code:
            continue

        panel = panel_cache.get(panel_code)
        if panel is None:
            panel = db.query(Panel).filter(Panel.project_id == project_id, Panel.panel_code == panel_code).first()
            if panel is None:
                panel = Panel(
                    project_id=project_id,
                    panel_code=panel_code,
                    panel_name=_string_value(row["panel_name"]) or panel_code,
                )
                db.add(panel)
                db.flush()
                summary["panels_created"] += 1
            else:
                summary["panels_updated"] += 1
            panel_cache[panel_code] = panel

        _update_panel_fields(
            panel,
            panel_name=_string_value(row["panel_name"]) or panel.panel_name or panel_code,
            serial_number=_string_value(row["panel_serial_number"]),
            location_note=_string_value(row["panel_location_note"]),
        )

        loop_key = (panel.id, loop_code)
        loop = loop_cache.get(loop_key)
        if loop is None:
            loop = db.query(Loop).filter(Loop.panel_id == panel.id, Loop.loop_code == loop_code).first()
            if loop is None:
                loop = Loop(
                    panel_id=panel.id,
                    loop_code=loop_code,
                    loop_name=_string_value(row["loop_name"]) or loop_code,
                )
                db.add(loop)
                db.flush()
                summary["loops_created"] += 1
            else:
                summary["loops_updated"] += 1
            loop_cache[loop_key] = loop

        _update_loop_fields(
            loop,
            loop_name=_string_value(row["loop_name"]) or loop.loop_name or loop_code,
            converter_name=_string_value(row["converter_name"]),
            converter_ip=_string_value(row["converter_ip"]),
            mac_address=_string_value(row["mac_address"]),
        )

        meter = db.query(Meter).filter(Meter.loop_id == loop.id, Meter.meter_code == meter_code).first()
        if meter is None:
            meter = Meter(
                loop_id=loop.id,
                meter_code=meter_code,
                meter_name=_string_value(row["meter_name"]) or meter_code,
            )
            db.add(meter)
            summary["meters_created"] += 1
        else:
            summary["meters_updated"] += 1

        _update_meter_fields(
            meter,
            meter_name=_string_value(row["meter_name"]) or meter.meter_name or meter_code,
            serial_number=_string_value(row["meter_serial_number"]),
            device_address=_string_value(row["device_address"]),
            model=_string_value(row["model"]),
            ct_ratio=_string_value(row["ct_ratio"]),
            baud_rate=_string_value(row["baud_rate"]),
            status=_string_value(row["status"]) or "Active",
        )

    db.commit()
    return summary


def _import_oakwood_loop_workbook(db: Session, workbook, project_id: int):
    summary = _empty_import_summary()
    panel_cache: dict[str, Panel] = {}
    loop_cache: dict[tuple[int, str], Loop] = {}

    loop_sheets = [workbook[name] for name in workbook.sheetnames if name.lower().startswith("loop")]
    if not loop_sheets:
        raise ValueError("No loop sheets were found in this workbook.")

    for sheet in loop_sheets:
        panel_name = _cell_text(sheet["O3"]) or _cell_text(sheet["T3"]) or sheet.title
        loop_name = _cell_text(sheet["T3"]) or sheet.title
        loop_code = sheet.title
        location_or_ip = _cell_text(sheet["C5"])
        system_name = _cell_text(sheet["C4"])
        mac_address = _cell_text(sheet["H5"])
        location_note = " | ".join(part for part in (system_name, location_or_ip) if part)

        panel_code = None
        for row_index in range(11, 43):
            row_panel_code = _cell_text(sheet[f"D{row_index}"])
            if row_panel_code:
                panel_code = row_panel_code
                break
        if not panel_code:
            panel_code = panel_name

        panel = panel_cache.get(panel_code)
        if panel is None:
            panel = db.query(Panel).filter(Panel.project_id == project_id, Panel.panel_code == panel_code).first()
            if panel is None:
                panel = Panel(
                    project_id=project_id,
                    panel_code=panel_code,
                    panel_name=panel_name or panel_code,
                )
                db.add(panel)
                db.flush()
                summary["panels_created"] += 1
            else:
                summary["panels_updated"] += 1
            panel_cache[panel_code] = panel

        _update_panel_fields(
            panel,
            panel_name=panel_name or panel.panel_name or panel_code,
            serial_number=panel.serial_number,
            location_note=location_note or panel.location_note,
        )

        loop_key = (panel.id, loop_code)
        loop = loop_cache.get(loop_key)
        if loop is None:
            loop = db.query(Loop).filter(Loop.panel_id == panel.id, Loop.loop_code == loop_code).first()
            if loop is None:
                loop = Loop(
                    panel_id=panel.id,
                    loop_code=loop_code,
                    loop_name=loop_name or loop_code,
                )
                db.add(loop)
                db.flush()
                summary["loops_created"] += 1
            else:
                summary["loops_updated"] += 1
            loop_cache[loop_key] = loop

        _update_loop_fields(
            loop,
            loop_name=loop_name or loop.loop_name or loop_code,
            converter_name=system_name,
            converter_ip=location_or_ip,
            mac_address=mac_address,
        )

        model_headers = {
            "K": _cell_text(sheet["K10"]),
            "L": _cell_text(sheet["L10"]),
            "M": _cell_text(sheet["M10"]),
            "N": _cell_text(sheet["N10"]),
        }

        for row_index in range(11, 43):
            meter_code = _cell_text(sheet[f"B{row_index}"])
            serial_number = _cell_text(sheet[f"C{row_index}"])
            if not meter_code and not serial_number:
                continue

            meter_code = meter_code or serial_number or f"{loop_code}-{row_index}"
            meter = db.query(Meter).filter(Meter.loop_id == loop.id, Meter.meter_code == meter_code).first()
            if meter is None and serial_number:
                meter = db.query(Meter).filter(Meter.loop_id == loop.id, Meter.serial_number == serial_number).first()

            if meter is None:
                meter = Meter(
                    loop_id=loop.id,
                    meter_code=meter_code,
                    meter_name=_cell_text(sheet[f"B{row_index}"]) or meter_code,
                )
                db.add(meter)
                summary["meters_created"] += 1
            else:
                summary["meters_updated"] += 1

            selected_model = _pick_oakwood_model(sheet, row_index, model_headers)
            status = _oakwood_status(sheet, row_index)
            meter_name = _cell_text(sheet[f"B{row_index}"]) or meter.meter_name or meter_code

            _update_meter_fields(
                meter,
                meter_name=meter_name,
                serial_number=serial_number,
                device_address=_cell_text(sheet[f"F{row_index}"]),
                model=selected_model or meter.model,
                ct_ratio=_cell_text(sheet[f"H{row_index}"]),
                baud_rate=_cell_text(sheet[f"G{row_index}"]),
                status=status or meter.status or "Active",
            )

    db.commit()
    return summary


def _is_oakwood_loop_workbook(workbook) -> bool:
    sheet_names = set(workbook.sheetnames)
    return "Cover" in sheet_names and any(name.lower().startswith("loop") for name in workbook.sheetnames)


def _resolve_project_workbook_path(project: Project) -> Path | None:
    from ..utils.project_files import project_upload_dir
    
    # 1. Try dedicated workbook path
    if project.project_workbook_file_path:
        relative = project.project_workbook_file_path.removeprefix("/uploads/")
        path = UPLOADS_DIR / relative
        if path.exists() and not _is_report_template(path): return path
        
    # 2. Scan the 'workbooks' directory for ANY .xlsx file
    workbook_dir = project_upload_dir(UPLOADS_DIR, project, "documents", "workbooks")
    if workbook_dir.exists():
        xlsx_files = list(workbook_dir.glob("*.xlsx"))
        if xlsx_files:
            # Prefer files that are NOT the report template
            for f in xlsx_files:
                if not _is_report_template(f): return f

    # 3. Try template path ONLY if it's NOT the standard report template
    if project.template_file_path:
        relative = project.template_file_path.removeprefix("/uploads/")
        path = UPLOADS_DIR / relative
        if path.exists() and not _is_report_template(path): return path

    # 4. Fallback to default path
    if DEFAULT_PROJECT_WORKBOOK_PATH.exists():
        return DEFAULT_PROJECT_WORKBOOK_PATH
        
    return None


def _is_report_template(path: Path) -> bool:
    """Returns True if the file looks like the empty 'Templat-Report.xlsx'"""
    if "templat-report" in path.name.lower(): return True
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True)
        sheet_names = set(wb.sheetnames)
        wb.close()

        # It's a template if it has the report pages AND no loop sheets
        has_report_pages = {"Page (4)", "Page (5)", "Cover"}.issubset(sheet_names)
        has_loop_sheets = any(name.lower().startswith("loop") for name in sheet_names)
        
        return has_report_pages and not has_loop_sheets
    except:
        return False


def _pick_oakwood_model(sheet, row_index: int, model_headers: dict[str, str | None]) -> str | None:
    for column in ("K", "L", "M", "N"):
        if _cell_text(sheet[f"{column}{row_index}"]).upper() == "P":
            return model_headers.get(column)
    return None


def _oakwood_status(sheet, row_index: int) -> str:
    if _cell_text(sheet[f"P{row_index}"]).upper() == "P":
        return "Offline"
    if _cell_text(sheet[f"O{row_index}"]).upper() == "P":
        return "Active"
    return "Active"


def get_assets_from_oakwood_workbook(workbook):
    loops_data = []
    loop_sheet_names = [name for name in workbook.sheetnames if name.lower().startswith("loop")]

    if not loop_sheet_names:
        raise ValueError("No loop sheets were found in this workbook.")

    for sheet_name in loop_sheet_names:
        try:
            sheet = workbook[sheet_name]
            loop_name = _cell_text(sheet["T3"]) or sheet_name
            meters_data = []
            
            # Use get() or default to empty string to avoid crashes if Row 10 is missing
            model_headers = {
                "K": _cell_text(sheet["K10"]),
                "L": _cell_text(sheet["L10"]),
                "M": _cell_text(sheet["M10"]),
                "N": _cell_text(sheet["N10"]),
            }

            for row_index in range(11, 100): # Scan more rows just in case
                meter_code = _cell_text(sheet[f"B{row_index}"])
                serial_number = _cell_text(sheet[f"C{row_index}"])
                
                # Stop if we hit a row with no meter code AND no serial
                if not meter_code and not serial_number:
                    # But check a few more rows just in case there's a gap
                    if row_index > 50: break 
                    continue

                meter_code = meter_code or serial_number or f"{loop_name}-{row_index}"
                selected_model = _pick_oakwood_model(sheet, row_index, model_headers)
                status = _oakwood_status(sheet, row_index)
                meter_name = _cell_text(sheet[f"B{row_index}"]) or meter_code

                meters_data.append({
                    "meter_code": meter_code,
                    "meter_name": meter_name,
                    "serial_number": serial_number,
                    "device_address": _cell_text(sheet[f"F{row_index}"]),
                    "model": selected_model,
                    "ct_ratio": _cell_text(sheet[f"H{row_index}"]),
                    "baud_rate": _cell_text(sheet[f"G{row_index}"]),
                    "status": status,
                })
            
            if meters_data:
                loops_data.append({
                    "loop_name": loop_name,
                    "meters": meters_data
                })
        except Exception as e:
            print(f"Error processing sheet {sheet_name}: {e}")
            continue # Skip failed sheets instead of crashing the whole thing

    return loops_data


def _update_panel_fields(panel: Panel, panel_name: str | None, serial_number: str | None, location_note: str | None):
    panel.panel_name = panel_name or panel.panel_name or panel.panel_code
    panel.serial_number = serial_number or panel.serial_number
    panel.location_note = location_note or panel.location_note


def _update_loop_fields(loop: Loop, loop_name: str | None, converter_name: str | None, converter_ip: str | None, mac_address: str | None):
    loop.loop_name = loop_name or loop.loop_name or loop.loop_code
    loop.converter_name = converter_name or loop.converter_name
    loop.converter_ip = converter_ip or loop.converter_ip
    loop.mac_address = mac_address or loop.mac_address


def _update_meter_fields(
    meter: Meter,
    meter_name: str | None,
    serial_number: str | None,
    device_address: str | None,
    model: str | None,
    ct_ratio: str | None,
    baud_rate: str | None,
    status: str | None,
):
    meter.meter_name = meter_name or meter.meter_name or meter.meter_code
    meter.serial_number = serial_number or meter.serial_number
    meter.device_address = device_address or meter.device_address
    meter.model = model or meter.model
    meter.ct_ratio = ct_ratio or meter.ct_ratio
    meter.baud_rate = baud_rate or meter.baud_rate
    meter.status = status or meter.status or "Active"


def _cell_text(cell) -> str:
    return _string_value(cell.value)


def _string_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _empty_import_summary() -> dict:
    return {
        "panels_created": 0,
        "panels_updated": 0,
        "loops_created": 0,
        "loops_updated": 0,
        "meters_created": 0,
        "meters_updated": 0,
    }

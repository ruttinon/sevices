from pathlib import Path

from openpyxl import load_workbook


def analyze_template_file(template_path: Path) -> dict:
    workbook = load_workbook(template_path)
    detected_layout = _detect_layout(workbook)

    sheets = []
    total_images = 0
    total_merged_ranges = 0
    total_formula_cells = 0

    for worksheet in workbook.worksheets:
        image_anchors = []
        for image in getattr(worksheet, "_images", []):
            anchor = getattr(image.anchor, "_from", None)
            if anchor is not None:
                image_anchors.append(
                    {
                        "row": anchor.row + 1,
                        "column": anchor.col + 1,
                        "cell": f"{_column_letter(anchor.col + 1)}{anchor.row + 1}",
                    }
                )

        non_empty_cells = 0
        formula_cells = 0
        placeholder_cells = []
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value in (None, ""):
                    continue
                non_empty_cells += 1
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_cells += 1
                if isinstance(cell.value, str) and any(token in cell.value for token in ("#REF!", "#VALUE!", "#NAME?", "#N/A")):
                    placeholder_cells.append({"cell": cell.coordinate, "value": cell.value})

        merged_ranges = [str(cell_range) for cell_range in worksheet.merged_cells.ranges]
        total_images += len(image_anchors)
        total_merged_ranges += len(merged_ranges)
        total_formula_cells += formula_cells

        sheets.append(
            {
                "name": worksheet.title,
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column,
                "non_empty_cells": non_empty_cells,
                "formula_cells": formula_cells,
                "image_count": len(image_anchors),
                "image_anchors": image_anchors,
                "merged_range_count": len(merged_ranges),
                "merged_range_preview": merged_ranges[:12],
                "placeholder_cells": placeholder_cells[:24],
            }
        )

    return {
        "template_name": template_path.name,
        "template_path": str(template_path),
        "detected_layout": detected_layout,
        "sheet_count": len(workbook.sheetnames),
        "summary": {
            "total_images": total_images,
            "total_merged_ranges": total_merged_ranges,
            "total_formula_cells": total_formula_cells,
        },
        "layout_notes": _build_layout_notes(detected_layout),
        "input_sections": _build_input_sections(detected_layout),
        "sheets": sheets,
    }


def _detect_layout(workbook) -> str:
    sheet_names = set(workbook.sheetnames)
    loop_sheets = [sheet_name for sheet_name in workbook.sheetnames if sheet_name.lower().startswith("loop")]
    if "Cover" in sheet_names and loop_sheets:
        return "oakwood_loop_workbook"
    if {"Cover", "Page (B)", "Page (1)", "Page (2)", "Page (4)", "Page (5)"}.issubset(sheet_names):
        return "avera_multi_page"
    return "generic_workbook"


def _build_layout_notes(layout_name: str) -> list[str]:
    if layout_name == "oakwood_loop_workbook":
        return [
            "This workbook is organized as one Cover sheet plus one sheet per loop such as Loop1, Loop2, and so on.",
            "The Cover sheet is a calculated summary sheet. Many cells already contain formulas and should not be overwritten blindly.",
            "Each Loop sheet is a table-first maintenance form. The blank rows under the existing meter list are continuation rows for more devices, not photo placeholders.",
            "The images already embedded in the workbook are logos. This Oakwood workbook does not contain dedicated service photo slots like the previous Avera multi-page template.",
            "Service No., Date, comment cells, and meter status columns are the main editable areas for save-as reports.",
        ]
    if layout_name == "avera_multi_page":
        return [
            "This workbook follows the previous Cover/Page detail layout with dedicated detail pages and image slots.",
            "The report generator can continue mapping headers, checklist rows, and image placeholders to the detail pages.",
        ]
    return [
        "This workbook does not match a known EMS template pattern yet. The admin should inspect the sheet structure before mapping values automatically.",
    ]


def _build_input_sections(layout_name: str) -> list[dict]:
    if layout_name == "oakwood_loop_workbook":
        return [
            {
                "section": "Cover Summary",
                "sheet": "Cover",
                "purpose": "Project-level overview and automatic totals",
                "fields": [
                    {"label": "Project / customer title", "cell": "F2"},
                    {"label": "Loop totals summary", "cell_range": "B16:U24"},
                    {"label": "Device total summary", "cell_range": "I32:U33"},
                ],
                "notes": [
                    "Most values in this sheet are formulas linked to Loop sheets.",
                    "Use this sheet as a calculated summary, not a free-form data entry page.",
                ],
            },
            {
                "section": "Loop Header",
                "sheet_pattern": "Loop*",
                "purpose": "Header information entered once per loop report",
                "fields": [
                    {"label": "Project", "cell": "C3"},
                    {"label": "Panel building", "cell": "O3"},
                    {"label": "Loop name", "cell": "T3"},
                    {"label": "System", "cell": "C4"},
                    {"label": "Service number", "cell": "O4"},
                    {"label": "Service / location", "cell": "C5"},
                    {"label": "MAC address", "cell": "H5"},
                    {"label": "Date", "cell": "O5"},
                ],
                "notes": [
                    "Service number and date are blank input cells and should be populated on every save-as report.",
                    "Project, panel, loop, location, and MAC should come from real project asset data.",
                ],
            },
            {
                "section": "Meter Rows",
                "sheet_pattern": "Loop*",
                "purpose": "Per-meter data table",
                "fields": [
                    {"label": "Meter rows", "cell_range": "A11:T42"},
                    {"label": "Summary formulas", "cell_range": "A43:T43"},
                ],
                "notes": [
                    "Rows 11 to 42 support up to 32 meter records per loop sheet.",
                    "Blank rows after the current devices are continuation slots for additional meters.",
                    "Comment cells are in column T. Pass and status markers are in columns O to R.",
                ],
            },
            {
                "section": "Images",
                "sheet_pattern": "Loop*",
                "purpose": "Existing embedded assets",
                "fields": [
                    {"label": "Logo anchors", "cell_range": "B1, L9, T1"},
                ],
                "notes": [
                    "These are existing logos, not service photo slots.",
                    "If the project requires photo evidence inside the workbook, a new photo sheet or extra layout block must be added.",
                ],
            },
        ]

    if layout_name == "avera_multi_page":
        return [
            {
                "section": "Detail Pages",
                "sheet_pattern": "Page (4)/(5)",
                "purpose": "Checklist and photo placement",
                "fields": [
                    {"label": "Checklist rows", "cell_range": "C22:T31"},
                    {"label": "Photo anchors", "cell_range": "B35, J35, R35"},
                ],
                "notes": [
                    "This layout supports embedded service photos in the workbook.",
                ],
            },
        ]

    return []


def _column_letter(column_index: int) -> str:
    result = ""
    current = column_index
    while current:
        current, remainder = divmod(current - 1, 26)
        result = chr(65 + remainder) + result
    return result


def extract_template_photo_captions(template_path: Path) -> list[str]:
    """Extract photo caption labels from template Excel file.
    
    Looks for text cells near photo areas (rows 46-48) that describe what photos should be taken.
    Common patterns in Thai templates:
    - รูปการทำความสะอาดมิเตอร์ (meter cleaning photo)
    - รูปหน้าจอมิเตอร์ (meter display photo)
    - รูปบัตรกำกับมิเตอร์ (meter label photo)
    """
    try:
        workbook = load_workbook(template_path, data_only=True)
    except Exception:
        return ["รูปที่ 1", "รูปที่ 2", "รูปที่ 3"]
    
    captions = []
    
    # Look in Page (4) and Page (5) for photo caption areas
    for sheet_name in ['Page (4)', 'Page (5)']:
        if sheet_name not in workbook.sheetnames:
            continue
            
        sheet = workbook[sheet_name]
        
        # Look for caption cells typically in row 46-48 near photo areas
        # Check common caption positions
        caption_cells = [
            ('B', 47), ('J', 47), ('R', 47),  # Common caption row
            ('B', 48), ('J', 48), ('R', 48),  # Alternative caption row
            ('C', 47), ('K', 47), ('S', 47),  # Slight offset
        ]
        
        sheet_captions = []
        for col, row in caption_cells:
            cell_value = sheet[f"{col}{row}"].value
            if cell_value and isinstance(cell_value, str):
                # Filter out non-caption text (long text, formulas, etc.)
                text = cell_value.strip()
                if 5 < len(text) < 100 and not text.startswith('='):
                    # Looks like a caption
                    sheet_captions.append(text)
        
        if sheet_captions:
            captions.extend(sheet_captions)
    
    workbook.close()
    
    # Remove duplicates while preserving order
    seen = set()
    unique_captions = []
    for c in captions:
        if c not in seen:
            seen.add(c)
            unique_captions.append(c)
    
    # Return up to 3 captions, fallback to defaults if not found
    if len(unique_captions) >= 3:
        return unique_captions[:3]
    elif len(unique_captions) > 0:
        # Fill remaining with defaults
        defaults = ["รูปการทำความสะอาดมิเตอร์", "รูปหน้าจอมิเตอร์", "รูปบัตรกำกับมิเตอร์"]
        for i in range(len(unique_captions), 3):
            if defaults[i] not in seen:
                unique_captions.append(defaults[i])
        return unique_captions[:3]
    else:
        return ["รูปการทำความสะอาดมิเตอร์", "รูปหน้าจอมิเตอร์", "รูปบัตรกำกับมิเตอร์"]


def extract_template_checklist_items(template_path: Path) -> list[dict]:
    """Extract default checklist items from template Excel file.
    
    Looks for checklist labels in rows 22-31 (checklist area).
    Returns list of items with section info.
    If no data found, returns 10 empty items for user to fill in UI.
    """
    try:
        workbook = load_workbook(template_path, data_only=True)
    except Exception:
        # Return 10 empty items if can't open file
        return [
            {"id": f"item_{row}", "label": "", "section": "การตรวจสอบมิเตอร์", "order": row - 22}
            for row in range(22, 32)
        ]
    
    items = []
    
    # Look in Page (4) for checklist items
    if 'Page (4)' in workbook.sheetnames:
        sheet = workbook['Page (4)']
        
        # Checklist rows are typically 22-31
        for row in range(22, 32):
            cell_value = sheet[f"C{row}"].value
            if cell_value and isinstance(cell_value, str):
                text = cell_value.strip()
                # Skip empty, formulas, and non-checklist text
                if text and len(text) > 3 and not text.startswith('=') and not text.startswith('หัวข้อ'):
                    items.append({
                        "id": f"item_{row}",
                        "label": text,
                        "section": "การตรวจสอบมิเตอร์",
                        "order": row - 22
                    })
    
    workbook.close()
    
    # If no items found, return 10 empty items so user can fill in UI
    if len(items) == 0:
        return [
            {"id": f"item_{row}", "label": "", "section": "การตรวจสอบมิเตอร์", "order": row - 22}
            for row in range(22, 32)
        ]
    
    return items

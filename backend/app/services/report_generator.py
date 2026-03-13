from datetime import datetime
from pathlib import Path
import re
import unicodedata

from sqlalchemy.orm import Session

from ..config import DEFAULT_REPORT_TEMPLATE_PATH, UPLOADS_DIR
from ..models import Report, ServiceJob, Project, Meter
from ..services.report_draft_store import load_report_draft
from ..utils.helpers import build_public_file_path, generate_qr_code
from ..utils.project_files import project_upload_dir
from ..schemas import report_schema

AVERA_TEMPLATE_SHEETS = {"Cover", "Page (B)", "Page (1)", "Page (2)", "Page (4)", "Page (5)"}
DEFAULT_CHECKLIST = [
    "Visual inspection of cabinet and meter enclosure",
    "Verify nameplate and serial number",
    "Check communication wiring and converter",
    "Confirm device address and baud rate",
    "Check meter display and status",
    "Review model and CT ratio",
    "Capture photo evidence",
    "Record final maintenance note",
]

# Default customizable checklist templates for different job types
DEFAULT_CHECKLIST_TEMPLATES = {
    "PM": {
        "job_type": "PM",
        "name": "Preventive Maintenance",
        "topics": [
            {"id": "pm1_0", "section": "1. ตรวจสอบทั่วไป", "label": "ตรวจสอบสายไฟและการต่อสาย", "order": 0},
            {"id": "pm1_1", "section": "1. ตรวจสอบทั่วไป", "label": "ตรวจสอบ LED indicator", "order": 1},
            {"id": "pm1_2", "section": "1. ตรวจสอบทั่วไป", "label": "ตรวจสอบ display", "order": 2},
            {"id": "pm1_3", "section": "1. ตรวจสอบทั่วไป", "label": "ตรวจสอบ housing / กล่อง", "order": 3},
            {"id": "pm2_0", "section": "2. Communication", "label": "ทดสอบ Modbus communication", "order": 4},
            {"id": "pm2_1", "section": "2. Communication", "label": "ตรวจสอบ Baud Rate", "order": 5},
            {"id": "pm2_2", "section": "2. Communication", "label": "ตรวจสอบ Device Address", "order": 6},
            {"id": "pm2_3", "section": "2. Communication", "label": "ทดสอบ data logging", "order": 7},
            {"id": "pm3_0", "section": "3. ค่ามิเตอร์", "label": "ตรวจสอบค่า Voltage", "order": 8},
            {"id": "pm3_1", "section": "3. ค่ามิเตอร์", "label": "ตรวจสอบค่า Current", "order": 9},
            {"id": "pm3_2", "section": "3. ค่ามิเตอร์", "label": "ตรวจสอบค่า Power Factor", "order": 10},
            {"id": "pm3_3", "section": "3. ค่ามิเตอร์", "label": "ตรวจสอบค่า Energy (kWh)", "order": 11},
            {"id": "pm3_4", "section": "3. ค่ามิเตอร์", "label": "เปรียบเทียบค่ากับ reference", "order": 12},
            {"id": "pm4_0", "section": "4. CT / PT", "label": "ตรวจสอบ CT Ratio", "order": 13},
            {"id": "pm4_1", "section": "4. CT / PT", "label": "ตรวจสอบการติดตั้ง CT", "order": 14},
            {"id": "pm4_2", "section": "4. CT / PT", "label": "ตรวจสอบ polarity", "order": 15},
        ],
        "is_custom": False,
    },
    "MA": {
        "job_type": "MA",
        "name": "Maintenance Agreement",
        "topics": [
            {"id": "ma1_0", "section": "1. ซ่อมบำรุง", "label": "ทำความสะอาดอุปกรณ์", "order": 0},
            {"id": "ma1_1", "section": "1. ซ่อมบำรุง", "label": "ขันน็อตให้แน่น", "order": 1},
            {"id": "ma1_2", "section": "1. ซ่อมบำรุง", "label": "เปลี่ยนชิ้นส่วนที่สึกหรอ", "order": 2},
            {"id": "ma1_3", "section": "1. ซ่อมบำรุง", "label": "ปรับค่า parameter", "order": 3},
            {"id": "ma2_0", "section": "2. ทดสอบหลังซ่อม", "label": "ทดสอบการทำงานหลังซ่อม", "order": 4},
            {"id": "ma2_1", "section": "2. ทดสอบหลังซ่อม", "label": "ตรวจสอบค่าการวัดหลังซ่อม", "order": 5},
            {"id": "ma2_2", "section": "2. ทดสอบหลังซ่อม", "label": "ตรวจสอบ communication", "order": 6},
            {"id": "ma3_0", "section": "3. บันทึกผล", "label": "บันทึกค่าก่อนซ่อม", "order": 7},
            {"id": "ma3_1", "section": "3. บันทึกผล", "label": "บันทึกค่าหลังซ่อม", "order": 8},
            {"id": "ma3_2", "section": "3. บันทึกผล", "label": "บันทึกรายการที่เปลี่ยน", "order": 9},
            {"id": "ma3_3", "section": "3. บันทึกผล", "label": "สรุปสาเหตุและการแก้ไข", "order": 10},
        ],
        "is_custom": False,
    },
    "IM": {
        "job_type": "IM",
        "name": "Installation & Commissioning",
        "topics": [
            {"id": "im1_0", "section": "1. ติดตั้ง", "label": "ติดตั้งมิเตอร์และ CT", "order": 0},
            {"id": "im1_1", "section": "1. ติดตั้ง", "label": "เชื่อมต่อสาย communication", "order": 1},
            {"id": "im1_2", "section": "1. ติดตั้ง", "label": "ตั้งค่า device address", "order": 2},
            {"id": "im2_0", "section": "2. ทดสอบ", "label": "ทดสอบการอ่านค่า", "order": 3},
            {"id": "im2_1", "section": "2. ทดสอบ", "label": "ทดสอบ communication", "order": 4},
            {"id": "im2_2", "section": "2. ทดสอบ", "label": "ตรวจสอบค่า voltage/current", "order": 5},
            {"id": "im3_0", "section": "3. มอบงาน", "label": "สอนการใช้งาน", "order": 6},
            {"id": "im3_1", "section": "3. มอบงาน", "label": "ส่งมอบเอกสาร", "order": 7},
        ],
        "is_custom": False,
    },
    "EM": {
        "job_type": "EM",
        "name": "Emergency Maintenance",
        "topics": [
            {"id": "em1_0", "section": "1. ตรวจสอบปัญหา", "label": "ตรวจสอบสาเหตุเบื้องต้น", "order": 0},
            {"id": "em1_1", "section": "1. ตรวจสอบปัญหา", "label": "บันทึกอาการที่พบ", "order": 1},
            {"id": "em2_0", "section": "2. แก้ไข", "label": "ดำเนินการซ่อมแซม", "order": 2},
            {"id": "em2_1", "section": "2. แก้ไข", "label": "เปลี่ยนอุปกรณ์ถ้าจำเป็น", "order": 3},
            {"id": "em3_0", "section": "3. ตรวจสอบหลังแก้ไข", "label": "ทดสอบการทำงาน", "order": 4},
            {"id": "em3_1", "section": "3. ตรวจสอบหลังแก้ไข", "label": "ยืนยันค่าการวัด", "order": 5},
        ],
        "is_custom": False,
    },
}

PREFERRED_AVERA_TEMPLATE_PATH = Path(__file__).resolve().parents[2] / "teamp" / "Templat-Report.xlsx"


def _absolute_path_from_public(public_path: str) -> Path:
    relative = public_path.removeprefix("/uploads/")
    return UPLOADS_DIR / relative


def _build_report_filename(service_job: ServiceJob, extension: str) -> str:
    # Use service_date instead of utcnow for YYYYMMDD suffix
    date_fragment = service_job.service_date.strftime("%Y%m%d")
    project_fragment = _slugify_filename_fragment(service_job.project.name)
    asset_fragment = _slugify_filename_fragment(_report_asset_label(service_job))
    service_fragment = _slugify_filename_fragment(service_job.service_type)
    
    stem = f"report_{project_fragment}_{asset_fragment}_{service_fragment}_{date_fragment}"
    return f"{stem}.{extension}"


def generate_service_reports(db: Session, service_job: ServiceJob):
    report_draft = load_report_draft(service_job)
    project_dir = project_upload_dir(UPLOADS_DIR, service_job.project, "reports")
    project_dir.mkdir(parents=True, exist_ok=True)
    
    pdf_filename = _build_report_filename(service_job, "pdf")
    excel_filename = _build_report_filename(service_job, "xlsx")

    pdf_path = project_dir / pdf_filename
    excel_path = project_dir / excel_filename

    _generate_pdf_report(service_job, pdf_path)
    _generate_excel_report(service_job, excel_path, report_draft)

    created_reports = []
    for path in (pdf_path, excel_path):
        report = Report(
            service_id=service_job.id,
            file_path=build_public_file_path(UPLOADS_DIR, path),
        )
        db.add(report)
        created_reports.append(report)

    db.commit()
    for report in created_reports:
        db.refresh(report)
    return created_reports


def generate_loop_report(db: Session, project: Project, request_data: report_schema.GenerateLoopReportRequest):
    """Generate loop report with date-based filename and append-only mode support."""
    from fastapi import HTTPException
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl is not installed. Cannot generate Excel reports.")

    template_path = _resolve_template_path_for_project(project)
    if not template_path or not template_path.exists():
        raise HTTPException(status_code=404, detail="Default report template not found.")

    # Get inspection date for filename
    inspection_date = request_data.inspection_date
    date_fragment = inspection_date.strftime("%Y%m%d")

    output_dir = project_upload_dir(UPLOADS_DIR, project, "reports")
    output_dir.mkdir(parents=True, exist_ok=True)

    # Build date-based filename: report_YYYYMMDD.xlsx
    report_type_fragment = _slugify_filename_fragment(request_data.report_type)
    filename = f"report_{date_fragment}_{report_type_fragment}.xlsx"
    output_path = output_dir / filename

    # Check if file exists (append mode)
    existing_data = {}
    if request_data.append_mode and output_path.exists():
        try:
            existing_wb = load_workbook(output_path)
            # Load existing meter data from the file
            existing_data = _extract_existing_report_data(existing_wb)
            existing_wb.close()
        except Exception:
            pass  # If extraction fails, we'll start fresh

    workbook = load_workbook(template_path)

    # Get checklist template (custom or default)
    template_config = request_data.custom_template
    if not template_config:
        template_config = DEFAULT_CHECKLIST_TEMPLATES.get(
            request_data.report_type,
            DEFAULT_CHECKLIST_TEMPLATES["PM"]  # Fallback to PM
        )

    # 1. Get all meters from the selected loops
    page_num = 3  # Starts after Cover, Page(B), Page(1), Page(2)
    total_meters = len(request_data.meter_data)
    total_pages = 2 + (total_meters * 2)  # Cover + TOC + 2 pages per meter

    # Merge with existing data if in append mode
    meter_data_list = _merge_meter_data(request_data.meter_data, existing_data)

    for idx, meter_data in enumerate(meter_data_list):
        meter = db.query(Meter).filter(Meter.meter_code == meter_data.meter_code).first()
        if not meter:
            meter = Meter(
                meter_code=meter_data.meter_code,
                meter_name=meter_data.meter_name or meter_data.meter_code,
            )

        # Use meter name for sheet naming (clean version)
        meter_name_clean = _slugify_filename_fragment(meter_data.meter_name or meter_data.meter_code)[:20]
        # Prefix with index to maintain order: 001_MeterName_P4
        sheet_prefix = f"{idx+1:03d}_{meter_name_clean}"

        # Copy Page (4) and Page (5) for each meter
        ws_page4 = workbook.copy_worksheet(workbook["Page (4)"])
        ws_page4.title = f"{sheet_prefix}_P4"
        ws_page5 = workbook.copy_worksheet(workbook["Page (5)"])
        ws_page5.title = f"{sheet_prefix}_P5"

        _populate_meter_detail_sheets(
            workbook=workbook,
            project=project,
            meter=meter,
            request_data=request_data,
            meter_data=meter_data,
            page_number_start=page_num,
            total_pages=total_pages,
            template_config=template_config,
        )
        page_num += 2

    # Remove original template sheets
    if "Page (4)" in workbook.sheetnames:
        del workbook["Page (4)"]
    if "Page (5)" in workbook.sheetnames:
        del workbook["Page (5)"]

    # Save the workbook
    workbook.save(output_path)

    # Create a record in the database
    report = Report(
        project_id=project.id,
        file_path=build_public_file_path(UPLOADS_DIR, output_path),
        report_date=inspection_date,
    )
    db.add(report)
    db.commit()
    db.refresh(report)
    return report


def _populate_meter_detail_sheets(
    workbook, project: Project, meter: Meter,
    request_data: report_schema.GenerateLoopReportRequest,
    meter_data: report_schema.MeterReportData,
    page_number_start: int, total_pages: int,
    template_config: dict | None = None,
):
    loop = meter.loop
    panel = loop.panel if loop else None
    customer = project.customer
    report_number = f"EMS-{request_data.report_type.upper()}-{datetime.now().strftime('%Y%m%d')}"

    # Get the newly created sheets for this meter (find by title pattern)
    # Sheets are named: 001_MeterName_P4, 001_MeterName_P5
    sheet_titles = [ws.title for ws in workbook.worksheets]
    meter_name_clean = _slugify_filename_fragment(meter_data.meter_name or meter_data.meter_code)[:20]
    
    # Find sheets that contain the meter name
    matching_sheets = [t for t in sheet_titles if meter_name_clean in t and ('_P4' in t or '_P5' in t)]
    if len(matching_sheets) >= 2:
        ws_page4 = workbook[matching_sheets[0]]  # P4 sheet
        ws_page5 = workbook[matching_sheets[1]]  # P5 sheet
    else:
        # Fallback to original method if sheets not found
        ws_page4 = workbook[f"{page_number_start:03d}_{meter_name_clean}_P4"]
        ws_page5 = workbook[f"{page_number_start:03d}_{meter_name_clean}_P5"]

    # Prepare checklist data - transform from API format to template format
    checklist_items = meter_data.checklist or []

    # Ensure mandatory remarks for Fail status
    validated_checklist = []
    for item in checklist_items:
        remark = item.remark or ""
        if item.status == "Fail" and not remark:
            remark = "[REQUIRED REMARK MISSING]"
        validated_checklist.append({
            "label": item.label,
            "status": item.status,
            "remark": remark,
            "pass": item.status == "Pass",
            "fail": item.status == "Fail",
        })

    # Split into two pages (max 10 items per page)
    checklist_page1 = validated_checklist[:10]
    checklist_page2 = validated_checklist[10:20]

    # Pad with empty items if needed
    while len(checklist_page1) < 10:
        checklist_page1.append({"label": "", "pass": False, "fail": False, "remark": ""})
    while len(checklist_page2) < 10:
        checklist_page2.append({"label": "", "pass": False, "fail": False, "remark": ""})

    # Prepare photos with titles
    photos = meter_data.photos or []
    photos_page1 = photos[:3]
    photos_page2 = photos[3:6]

    # Common data for both pages
    common_data = {
        "panel": panel,
        "loop": loop,
        "meter": meter,
        "customer_name": customer.name,
        "project_name": project.name,
        "report_number": report_number,
        "service_date": request_data.inspection_date,
        "total_pages": total_pages,
        "overview_note": meter_data.comment or "",
        "meter_data": meter_data,
    }

    # Populate Page 4
    _populate_detail_sheet_v2(
        sheet=ws_page4,
        page_number=page_number_start,
        checklist_rows=checklist_page1,
        photos=photos_page1,
        **common_data,
    )

    # Populate Page 5
    _populate_detail_sheet_v2(
        sheet=ws_page5,
        page_number=page_number_start + 1,
        checklist_rows=checklist_page2,
        photos=photos_page2,
        **common_data,
    )


def _populate_detail_sheet_v2(
    sheet,
    panel,
    loop,
    meter,
    customer_name: str,
    project_name: str,
    report_number: str,
    service_date,
    page_number: int,
    total_pages: int,
    checklist_rows: list[dict],
    photos: list[report_schema.PhotoWithTitle],
    overview_note: str,
    meter_data: report_schema.MeterReportData,
):
    """Populate a detail sheet (Page 4 or 5) with full data including photo titles."""
    # Header info
    sheet["G5"] = project_name
    sheet["G6"] = customer_name
    sheet["G7"] = _format_service_date(service_date)
    sheet["Q2"] = report_number
    sheet["V2"] = page_number
    sheet["W53"] = page_number
    sheet["Y53"] = total_pages

    # Meter detail section
    sheet["G11"] = project_name
    sheet["U11"] = _panel_label(panel)
    sheet["G12"] = customer_name
    sheet["U12"] = panel.location_note if panel and panel.location_note else "-"

    # Main meter data fields
    sheet["E15"] = meter_data.meter_name or meter.meter_name or "-"
    sheet["P15"] = meter_data.meter_code or "-"
    sheet["E16"] = meter_data.online_status if meter_data.online_status is not None else (meter.serial_number or "-")
    sheet["P16"] = meter.device_address if meter and meter.device_address else "-"
    sheet["E17"] = panel.location_note if panel and panel.location_note else "-"
    sheet["P17"] = loop.loop_name if loop else "-"
    sheet["E18"] = meter.ct_ratio if meter and meter.ct_ratio else "-"

    # Populate checklist rows (rows 22-31 in template)
    for idx, item in enumerate(checklist_rows[:10]):
        row = 22 + idx
        sheet[f"C{row}"] = item["label"]
        # Use symbols instead of boolean TRUE/FALSE
        if item["pass"]:
            sheet[f"P{row}"] = "✓"  # Pass checkmark
        else:
            sheet[f"P{row}"] = ""  # Empty for not pass
        
        if item["fail"]:
            sheet[f"R{row}"] = "✗"  # Fail X mark
        else:
            sheet[f"R{row}"] = ""  # Empty for not fail
        
        sheet[f"T{row}"] = item["remark"]

    # Insert photos with titles
    _insert_template_photos_v2(sheet, photos)

    # Overall note
    if overview_note:
        sheet["T40"] = overview_note


def _insert_template_photos_v2(sheet, photos: list[report_schema.PhotoWithTitle]):
    """Insert photos with editable titles into template."""
    photo_slots = [("B35", "B46", 215, 135), ("J35", "J46", 215, 135), ("R35", "R46", 215, 135)]

    try:
        from openpyxl.drawing.image import Image as OpenPyxlImage
    except ImportError:
        # Fallback: just put file paths
        for (cell, title_cell, _, _), photo in zip(photo_slots, photos):
            if photo and photo.file_path:
                sheet[cell] = photo.file_path
                if photo.title:
                    sheet[title_cell] = photo.title
        return

    for (cell, title_cell, width, height), photo in zip(photo_slots, photos):
        if not photo or not photo.file_path:
            continue

        image_path = _absolute_path_from_public(photo.file_path)
        if not image_path.exists():
            sheet[cell] = photo.file_path
            if photo.title:
                sheet[title_cell] = photo.title
            continue

        try:
            image = OpenPyxlImage(str(image_path))
            image.width = width
            image.height = height
            sheet.add_image(image, cell)
        except Exception:
            sheet[cell] = image_path.name

        # Add photo title
        if photo.title:
            sheet[title_cell] = photo.title
        elif photo.caption:
            sheet[title_cell] = photo.caption


def _extract_existing_report_data(workbook) -> dict:
    """Extract existing meter data from a report file for append mode."""
    existing_data = {}
    for sheet_name in workbook.sheetnames:
        if sheet_name.startswith("Meter_") and sheet_name.endswith("_P4"):
            # Extract meter code from sheet name
            meter_code = sheet_name.replace("Meter_", "").replace("_P4", "")
            sheet = workbook[sheet_name]

            # Extract meter info from the sheet
            meter_name = sheet["E15"].value
            serial = sheet["E16"].value
            comment = sheet["T40"].value

            # Extract checklist items
            checklist = []
            for row in range(22, 32):
                label = sheet[f"C{row}"].value
                if label:
                    pass_val = sheet[f"P{row}"].value
                    fail_val = sheet[f"R{row}"].value
                    remark = sheet[f"T{row}"].value
                    status = "Pass" if pass_val else ("Fail" if fail_val else "N/A")
                    checklist.append({
                        "label": label,
                        "status": status,
                        "remark": remark or "",
                    })

            existing_data[meter_code] = {
                "meter_name": meter_name,
                "serial_number": serial,
                "comment": comment,
                "checklist": checklist,
            }

    return existing_data


def _merge_meter_data(
    new_data: list[report_schema.MeterReportData],
    existing_data: dict
) -> list[report_schema.MeterReportData]:
    """Merge new meter data with existing data for append mode."""
    # Create a map of new data by meter code
    new_by_code = {m.meter_code: m for m in new_data}

    # Start with all new data
    merged = list(new_data)

    # Add any meters from existing data that aren't in new data
    for meter_code, existing in existing_data.items():
        if meter_code not in new_by_code:
            # Create a MeterReportData from existing
            merged.append(report_schema.MeterReportData(
                meter_code=meter_code,
                meter_name=existing.get("meter_name"),
                online_status=True,
                accuracy_status="Pass",
                energy_reading=None,
                comment=existing.get("comment"),
                checklist=[report_schema.MeterChecklistItem(
                    label=c["label"],
                    status=c["status"],
                    remark=c["remark"],
                ) for c in existing.get("checklist", [])],
                photos=[],
            ))

    return merged


def get_checklist_template(job_type: str) -> dict:
    """Get the default checklist template for a job type."""
    return DEFAULT_CHECKLIST_TEMPLATES.get(
        job_type.upper(),
        DEFAULT_CHECKLIST_TEMPLATES["PM"]  # Fallback
    )


def _resolve_template_path_for_project(project: Project) -> Path | None:
    if project.template_file_path:
        candidate_path = _absolute_path_from_public(project.template_file_path)
        if candidate_path.exists():
            return candidate_path

    if PREFERRED_AVERA_TEMPLATE_PATH.exists():
        return PREFERRED_AVERA_TEMPLATE_PATH

    return DEFAULT_REPORT_TEMPLATE_PATH if DEFAULT_REPORT_TEMPLATE_PATH.exists() else None


def _generate_pdf_report(service_job: ServiceJob, output_path: Path):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError:
        output_path.write_text(
            "\n".join(
                [
                    "EMS Service Report",
                    f"Project: {service_job.project.name}",
                    f"Customer: {service_job.project.customer.name}",
                    f"Service Type: {service_job.service_type}",
                    f"Service Date: {service_job.service_date.isoformat()}",
                    f"Engineer: {service_job.engineer_name or service_job.engineer_id or '-'}",
                    f"Status: {service_job.status}",
                    f"Note: {service_job.note or '-'}",
                ]
            ),
            encoding="utf-8",
        )
        return

    styles = getSampleStyleSheet()
    story = []

    project = service_job.project
    meter = service_job.meter
    loop = meter.loop if meter else None
    panel = loop.panel if loop else None
    customer = project.customer

    story.append(Paragraph("EMS Service Report", styles["Title"]))
    story.append(Spacer(1, 0.2 * inch))

    summary_rows = [
        ["Customer", customer.name],
        ["Project", project.name],
        ["Location", project.location],
        ["Service Type", service_job.service_type],
        ["Service Date", service_job.service_date.strftime("%Y-%m-%d %H:%M")],
        ["Engineer", service_job.engineer_name or str(service_job.engineer_id or "-")],
        ["Status", service_job.status],
    ]
    if meter:
        summary_rows.extend(
            [
                ["Panel", panel.panel_name if panel else "-"],
                ["Loop", loop.loop_name if loop else "-"],
                ["Meter", meter.meter_name],
                ["Meter Serial", meter.serial_number or "-"],
            ]
        )

    table = Table(summary_rows, hAlign="LEFT", colWidths=[1.8 * inch, 4.5 * inch])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#12344d")),
                ("TEXTCOLOR", (0, 0), (0, -1), colors.white),
                ("BACKGROUND", (1, 0), (1, -1), colors.HexColor("#f8fbfd")),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d4dde6")),
                ("PADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(table)
    story.append(Spacer(1, 0.2 * inch))
    story.append(Paragraph(f"Service Note: {service_job.note or '-'}", styles["BodyText"]))
    story.append(Spacer(1, 0.2 * inch))

    checklist_rows = _normalize_checklist_rows(service_job)
    if checklist_rows:
        story.append(Paragraph("Maintenance Checklist", styles["Heading2"]))
        checklist_table = Table(
            [["Item", "Status", "Remark"]] + [
                [item["label"], "Pass" if item["pass"] else "Fail", item["remark"] or "-"]
                for item in checklist_rows
            ],
            hAlign="LEFT",
            colWidths=[3.6 * inch, 0.9 * inch, 2.0 * inch],
        )
        checklist_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#12344d")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#f8fbfd")),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d4dde6")),
                    ("PADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        story.append(checklist_table)
        story.append(Spacer(1, 0.2 * inch))

    if service_job.photos:
        story.append(Paragraph("Site Photos", styles["Heading2"]))
        for photo in service_job.photos[:3]:
            image_path = _absolute_path_from_public(photo.file_path)
            if image_path.exists():
                story.append(Image(str(image_path), width=2.6 * inch, height=2.0 * inch))
                if photo.caption:
                    story.append(Paragraph(photo.caption, styles["BodyText"]))
                story.append(Spacer(1, 0.1 * inch))

    doc = SimpleDocTemplate(str(output_path), pagesize=A4)
    doc.build(story)


def _generate_excel_report(service_job: ServiceJob, output_path: Path, report_draft: dict | None = None):
    template_path = _resolve_template_path(service_job)

    if template_path is not None:
        _generate_excel_from_template(service_job, template_path, output_path, report_draft)
        return

    try:
        import pandas as pd
    except ImportError:
        output_path.write_text(
            "field,value\n"
            f"customer,{service_job.project.customer.name}\n"
            f"project,{service_job.project.name}\n"
            f"service_type,{service_job.service_type}\n"
            f"service_date,{service_job.service_date.isoformat()}\n"
            f"status,{service_job.status}\n",
            encoding="utf-8",
        )
        return

    meter = service_job.meter
    loop = meter.loop if meter else None
    panel = loop.panel if loop else None
    project = service_job.project
    customer = project.customer
    report_draft = report_draft or {}

    summary = pd.DataFrame(
        [
            {"field": "customer", "value": customer.name},
            {"field": "project", "value": project.name},
            {"field": "location", "value": project.location},
            {"field": "service_type", "value": service_job.service_type},
            {"field": "service_date", "value": service_job.service_date.isoformat()},
            {"field": "engineer_name", "value": service_job.engineer_name or service_job.engineer_id},
            {"field": "inspection_product", "value": _draft_value(report_draft, "inspection_product", _build_inspection_product(service_job))},
            {"field": "approve_by", "value": _draft_value(report_draft, "approve_by", customer.contact_name or "EMS Platform")},
            {"field": "status", "value": service_job.status},
            {"field": "note", "value": service_job.note},
            {"field": "panel", "value": panel.panel_name if panel else None},
            {"field": "loop", "value": loop.loop_name if loop else None},
            {"field": "meter", "value": meter.meter_name if meter else None},
            {"field": "meter_serial", "value": meter.serial_number if meter else None},
        ]
    )
    photos = pd.DataFrame(
        [{"file_path": photo.file_path, "caption": photo.caption, "created_at": photo.created_at.isoformat()} for photo in service_job.photos]
    )
    checklist = pd.DataFrame(_checklist_sheet_rows(service_job))

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="summary", index=False)
        checklist.to_excel(writer, sheet_name="checklist", index=False)
        photos.to_excel(writer, sheet_name="photos", index=False)


def _resolve_template_path(service_job: ServiceJob) -> Path | None:
    if service_job.project.template_file_path:
        candidate_path = _absolute_path_from_public(service_job.project.template_file_path)
        if candidate_path.exists() and _workbook_matches_avera_template(candidate_path):
            return candidate_path

    if PREFERRED_AVERA_TEMPLATE_PATH.exists():
        return PREFERRED_AVERA_TEMPLATE_PATH

    if DEFAULT_REPORT_TEMPLATE_PATH.exists():
        return DEFAULT_REPORT_TEMPLATE_PATH

    if service_job.project.template_file_path:
        candidate_path = _absolute_path_from_public(service_job.project.template_file_path)
        if candidate_path.exists():
            return candidate_path

    return None


def _generate_excel_from_template(service_job: ServiceJob, template_path: Path, output_path: Path, report_draft: dict | None = None):
    try:
        from openpyxl import load_workbook
    except ImportError:
        template_bytes = template_path.read_bytes()
        output_path.write_bytes(template_bytes)
        return

    workbook = load_workbook(template_path)
    report_draft = report_draft or {}

    if _is_avera_template(workbook):
        _populate_avera_template(workbook, service_job, report_draft)
        workbook.save(output_path)
        return

    meter = service_job.meter
    loop = meter.loop if meter else None
    panel = loop.panel if loop else None
    project = service_job.project
    customer = project.customer

    for sheet_name in ("EMS_SUMMARY", "EMS_EQUIPMENT", "EMS_CHECKLIST", "EMS_PHOTOS"):
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            worksheet.delete_rows(1, worksheet.max_row)
        else:
            worksheet = workbook.create_sheet(sheet_name)

        if sheet_name == "EMS_SUMMARY":
            worksheet.append(["field", "value"])
            worksheet.append(["customer", customer.name])
            worksheet.append(["project", project.name])
            worksheet.append(["location", project.location])
            worksheet.append(["service_type", service_job.service_type])
            worksheet.append(["service_date", service_job.service_date.isoformat()])
            worksheet.append(["engineer", service_job.engineer_name or service_job.engineer_id or "-"])
            worksheet.append(["status", service_job.status])
            worksheet.append(["note", service_job.note or "-"])
        elif sheet_name == "EMS_EQUIPMENT":
            worksheet.append(["panel_code", "panel_name", "loop_code", "loop_name", "meter_code", "meter_name", "serial_number", "model", "ct_ratio", "baud_rate", "status"])
            if meter is not None:
                worksheet.append([
                    panel.panel_code if panel else "-",
                    panel.panel_name if panel else "-",
                    loop.loop_code if loop else "-",
                    loop.loop_name if loop else "-",
                    meter.meter_code,
                    meter.meter_name,
                    meter.serial_number or "-",
                    meter.model or "-",
                    meter.ct_ratio or "-",
                    meter.baud_rate or "-",
                    meter.status,
                ])
        elif sheet_name == "EMS_CHECKLIST":
            worksheet.append(["label", "status", "note"])
            for item in _normalize_checklist_rows(service_job):
                worksheet.append([item["label"], "Pass" if item["pass"] else "Fail", item["remark"]])
        elif sheet_name == "EMS_PHOTOS":
            worksheet.append(["file_path", "caption", "created_at"])
            for photo in service_job.photos:
                worksheet.append([photo.file_path, photo.caption or "", photo.created_at.isoformat()])

    workbook.save(output_path)


def _workbook_matches_avera_template(template_path: Path) -> bool:
    try:
        from openpyxl import load_workbook
        workbook = load_workbook(template_path, read_only=True)
    except Exception:
        return False
    return _is_avera_template(workbook)


def _is_avera_template(workbook) -> bool:
    return AVERA_TEMPLATE_SHEETS.issubset(set(workbook.sheetnames))


def _populate_avera_template(workbook, service_job: ServiceJob, report_draft: dict | None = None):
    meter = service_job.meter
    loop = meter.loop if meter else None
    panel = loop.panel if loop else None
    project = service_job.project
    customer = project.customer
    report_draft = report_draft or {}
    inspection_product = _draft_value(report_draft, "inspection_product", _build_inspection_product(service_job))
    inspection_by = _draft_value(report_draft, "inspection_by", service_job.engineer_name or service_job.engineer_id or "Field Engineer")
    approve_by = _draft_value(report_draft, "approve_by", customer.contact_name or "EMS Platform")
    overview_note = _draft_value(report_draft, "overview_note", service_job.note or "No additional note recorded.")

    equipment_rows = _collect_equipment_rows(service_job)
    report_number = _build_report_number(service_job)
    total_pages = 4

    cover = workbook["Cover"]
    cover["I21"] = project.name
    cover["I22"] = _format_service_date(service_job.service_date)
    cover["I23"] = inspection_product
    cover["I24"] = customer.name
    cover["H26"] = inspection_by
    cover["H28"] = approve_by

    toc_sheet = workbook["Page (B)"]
    _populate_common_template_header(
        toc_sheet,
        project_name=project.name,
        customer_name=customer.name,
        report_number=report_number,
        service_date=service_job.service_date,
    )
    toc_sheet["D12"] = "Introduction"
    toc_sheet["D13"] = "Equipment list"
    toc_sheet["D14"] = "Inspection detail"
    toc_sheet["D15"] = "Inspection photos"
    toc_sheet["W40"] = 0
    toc_sheet["Y40"] = total_pages

    intro_sheet = workbook["Page (1)"]
    _populate_common_template_header(
        intro_sheet,
        project_name=project.name,
        customer_name=customer.name,
        report_number=report_number,
        service_date=service_job.service_date,
    )
    intro_sheet["W51"] = 1
    intro_sheet["Y51"] = total_pages
    intro_sheet["C12"] = (
        f"This report records the {inspection_product} for {project.name}."
    )
    intro_sheet["B13"] = (
        f"Customer: {customer.name} | Location: {project.location}"
    )
    intro_sheet["B15"] = (
        f"Panel: {_panel_label(panel)} | Loop: {loop.loop_name if loop else '-'}"
    )
    intro_sheet["B16"] = (
        f"Meter: {meter.meter_name if meter else '-'} | Meter code: {meter.meter_code if meter else '-'}"
    )
    intro_sheet["B17"] = (
        f"Serial: {meter.serial_number if meter and meter.serial_number else '-'} | Model: {meter.model if meter and meter.model else '-'}"
    )
    intro_sheet["B18"] = (
        f"Engineer: {inspection_by or '-'} | Status: {service_job.status}"
    )
    intro_sheet["B19"] = (
        f"Service date: {_format_service_datetime(service_job.service_date)} | Service note: {overview_note}"
    )
    intro_sheet["B21"] = (
        "The generated workbook is created as a save-as copy. The original template remains unchanged."
    )

    equipment_sheet = workbook["Page (2)"]
    _populate_common_template_header(
        equipment_sheet,
        project_name=project.name,
        customer_name=customer.name,
        report_number=report_number,
        service_date=service_job.service_date,
    )
    equipment_sheet["W51"] = 2
    equipment_sheet["Y51"] = total_pages
    equipment_sheet["C12"] = "No."
    equipment_sheet["F12"] = "Meter Name"
    equipment_sheet["K12"] = "Serial Number"
    equipment_sheet["P12"] = "Panel"
    equipment_sheet["T12"] = "Loop"
    equipment_sheet["X12"] = "Status"
    equipment_sheet["Y12"] = "Model"

    for row_index in range(13, 45):
        for column in ("C", "F", "K", "P", "T", "X", "Y"):
            equipment_sheet[f"{column}{row_index}"] = None

    for row_offset, equipment in enumerate(equipment_rows[:32], start=13):
        equipment_sheet[f"C{row_offset}"] = row_offset - 12
        equipment_sheet[f"F{row_offset}"] = equipment["meter_name"]
        equipment_sheet[f"K{row_offset}"] = equipment["serial_number"]
        equipment_sheet[f"P{row_offset}"] = equipment["panel_code"]
        equipment_sheet[f"T{row_offset}"] = equipment["loop_name"]
        equipment_sheet[f"X{row_offset}"] = equipment["status"]
        equipment_sheet[f"Y{row_offset}"] = equipment["model"]

    checklist_rows = _normalize_checklist_rows(service_job)
    detail_page_one = workbook["Page (4)"]
    detail_page_two = workbook["Page (5)"]
    _populate_detail_sheet(
        detail_page_one,
        service_job=service_job,
        panel=panel,
        loop=loop,
        meter=meter,
        customer_name=customer.name,
        project_name=project.name,
        report_number=report_number,
        service_date=service_job.service_date,
        page_number=3,
        total_pages=total_pages,
        checklist_rows=checklist_rows[:10],
        photos=service_job.photos[:3],
        overview_note=overview_note,
    )
    _populate_detail_sheet(
        detail_page_two,
        service_job=service_job,
        panel=panel,
        loop=loop,
        meter=meter,
        customer_name=customer.name,
        project_name=project.name,
        report_number=report_number,
        service_date=service_job.service_date,
        page_number=4,
        total_pages=total_pages,
        checklist_rows=checklist_rows[10:20],
        photos=service_job.photos[3:6],
        overview_note=overview_note,
    )


def _populate_common_template_header(sheet, project_name: str, customer_name: str, report_number: str, service_date):
    sheet["G5"] = project_name
    sheet["G6"] = customer_name
    sheet["G7"] = _format_service_date(service_date)
    sheet["Q2"] = report_number


def _populate_detail_sheet(
    sheet,
    service_job: ServiceJob,
    panel,
    loop,
    meter,
    customer_name: str,
    project_name: str,
    report_number: str,
    service_date,
    page_number: int,
    total_pages: int,
    checklist_rows,
    photos,
    overview_note: str,
):
    _populate_common_template_header(
        sheet,
        project_name=project_name,
        customer_name=customer_name,
        report_number=report_number,
        service_date=service_date,
    )
    sheet["V2"] = page_number
    sheet["G11"] = project_name
    sheet["U11"] = _panel_label(panel)
    sheet["G12"] = customer_name
    sheet["U12"] = panel.location_note if panel and panel.location_note else service_job.project.location
    sheet["E15"] = meter.meter_name if meter else "-"
    sheet["P15"] = meter.meter_code if meter else "-"
    sheet["E16"] = meter.serial_number if meter and meter.serial_number else "-"
    sheet["P16"] = meter.device_address if meter and meter.device_address else "-"
    sheet["E17"] = panel.location_note if panel and panel.location_note else service_job.project.location
    sheet["P17"] = loop.loop_name if loop else "-"
    sheet["E18"] = meter.ct_ratio if meter and meter.ct_ratio else "-"
    sheet["W53"] = page_number
    sheet["Y53"] = total_pages

    for row_number in range(22, 32):
        sheet[f"C{row_number}"] = ""
        sheet[f"P{row_number}"] = False
        sheet[f"R{row_number}"] = False
        sheet[f"T{row_number}"] = ""

    for row_number, item in zip(range(22, 32), checklist_rows):
        sheet[f"C{row_number}"] = item["label"]
        sheet[f"P{row_number}"] = item["pass"]
        sheet[f"R{row_number}"] = item["fail"]
        sheet[f"T{row_number}"] = item["remark"]
        
        # Validation: Require remark if fail
        if item["fail"] and not item["remark"]:
            sheet[f"T{row_number}"] = "⚠️ [REQUIRED REMARK MISSING]"

    if not checklist_rows and overview_note:
        sheet["T22"] = overview_note

    _insert_template_photos(sheet, photos)


def _collect_equipment_rows(service_job: ServiceJob) -> list[dict]:
    meter = service_job.meter
    if meter and meter.loop:
        meters = sorted(meter.loop.meters, key=lambda item: str(item.meter_code or item.id))
    elif service_job.project.panels:
        meters = []
        for panel in service_job.project.panels:
            for loop in panel.loops:
                meters.extend(loop.meters)
        meters = sorted(meters, key=lambda item: str(item.meter_code or item.id))
    else:
        meters = [meter] if meter else []

    rows = []
    for current_meter in meters:
        current_loop = current_meter.loop
        current_panel = current_loop.panel if current_loop else None
        rows.append(
            {
                "meter_name": current_meter.meter_name,
                "serial_number": current_meter.serial_number or "-",
                "panel_code": current_panel.panel_code if current_panel else "-",
                "loop_name": current_loop.loop_name if current_loop else "-",
                "status": current_meter.status,
                "model": current_meter.model or "-",
            }
        )
    return rows


def _normalize_checklist_rows(service_job: ServiceJob) -> list[dict]:
    meter = service_job.meter
    loop = meter.loop if meter else None
    panel = loop.panel if loop else None
    completed = str(service_job.status).lower() == "completed"
    engineer = str(service_job.engineer_name or service_job.engineer_id or "-")
    detail_remarks = {
        0: _panel_label(panel),
        1: meter.serial_number if meter and meter.serial_number else "-",
        2: _network_label(loop),
        3: _communication_label(meter, loop),
        4: meter.status if meter else service_job.status,
        5: f"{meter.model if meter and meter.model else '-'} / {meter.ct_ratio if meter and meter.ct_ratio else '-'}",
        6: ", ".join(photo.caption for photo in service_job.photos if photo.caption) or "Photo and service note recorded in EMS Platform.",
        7: service_job.note or f"Completed by {engineer}",
    }

    rows = []
    source_items = service_job.checklist_items or []
    labels = DEFAULT_CHECKLIST if not source_items else [item.get("label") or DEFAULT_CHECKLIST[min(index, len(DEFAULT_CHECKLIST) - 1)] for index, item in enumerate(source_items)]
    for index, label in enumerate(labels):
        item = source_items[index] if index < len(source_items) else {}
        item_status = str(item.get("status", "pass")).lower()
        item_note = item.get("note")
        rows.append(
            {
                "label": label,
                "pass": item_status not in {"fail", "not_pass", "not-pass", "offline"} if source_items else completed,
                "fail": item_status in {"fail", "not_pass", "not-pass", "offline"} if source_items else False,
                "remark": item_note or detail_remarks.get(index, ""),
            }
        )
    return rows


def _checklist_sheet_rows(service_job: ServiceJob) -> list[dict]:
    return [
        {
            "label": item["label"],
            "status": "Pass" if item["pass"] else "Fail",
            "note": item["remark"],
        }
        for item in _normalize_checklist_rows(service_job)
    ]


def _insert_template_photos(sheet, photos):
    photo_slots = [("B35", 215, 135), ("J35", 215, 135), ("R35", 215, 135)]

    try:
        from openpyxl.drawing.image import Image as OpenPyxlImage
    except ImportError:
        for (cell, _, _), photo in zip(photo_slots, photos):
            sheet[cell] = photo.file_path
        return

    for (cell, width, height), photo in zip(photo_slots, photos):
        image_path = _absolute_path_from_public(photo.file_path)
        if not image_path.exists():
            sheet[cell] = photo.file_path
            continue
        try:
            image = OpenPyxlImage(str(image_path))
            image.width = width
            image.height = height
            sheet.add_image(image, cell)
        except Exception:
            sheet[cell] = image_path.name
        caption_row = int(cell[1:]) + 11
        caption_cell = f"{cell[0]}{caption_row}"
        if photo.caption:
            sheet[caption_cell] = photo.caption


def _draft_value(report_draft: dict | None, key: str, fallback):
    value = (report_draft or {}).get(key)
    if value in (None, ""):
        return fallback
    return value


def _build_report_number(service_job: ServiceJob) -> str:
    return f"EMS-{service_job.service_type.upper()}-{service_job.id:04d}"


def _report_asset_label(service_job: ServiceJob) -> str:
    meter = service_job.meter
    loop = meter.loop if meter else None
    panel = loop.panel if loop else None
    if panel and panel.panel_code:
        return panel.panel_code
    if loop and loop.loop_code:
        return loop.loop_code
    if meter and meter.meter_code:
        return meter.meter_code
    return service_job.project.name


def _slugify_filename_fragment(value: str | None) -> str:
    normalized = unicodedata.normalize("NFKD", value or "")
    ascii_value = normalized.encode("ascii", "ignore").decode("ascii")
    slug = re.sub(r"[^A-Za-z0-9]+", "-", ascii_value).strip("-").lower()
    return slug or generate_qr_code("RPT")[-6:].lower()


def _sanitize_filename_fragment(value: str | None) -> str:
    invalid_chars = '<>:"/\\|?*'
    cleaned = "".join("-" if char in invalid_chars else char for char in str(value or "")).strip().strip(".")
    return cleaned[:80]


def _build_inspection_product(service_job: ServiceJob) -> str:
    meter = service_job.meter
    loop = meter.loop if meter else None
    panel = loop.panel if loop else None
    target_name = _panel_label(panel) if panel else (meter.meter_name if meter else service_job.project.name)
    return f"{service_job.service_type.upper()} service report for {target_name}"


def _panel_label(panel) -> str:
    if panel is None:
        return "-"
    if panel.panel_name and panel.panel_code:
        return f"{panel.panel_code} / {panel.panel_name}"
    return panel.panel_name or panel.panel_code or "-"


def _communication_label(meter, loop) -> str:
    if meter is None:
        return "-"
    parts = []
    if meter.device_address:
        parts.append(f"Address {meter.device_address}")
    if meter.baud_rate:
        parts.append(f"Baud {meter.baud_rate}")
    if loop and loop.converter_ip:
        parts.append(f"IP {loop.converter_ip}")
    return " | ".join(parts) if parts else "-"


def _network_label(loop) -> str:
    if loop is None:
        return "-"
    parts = [loop.loop_name]
    if loop.converter_ip:
        parts.append(loop.converter_ip)
    if loop.mac_address:
        parts.append(loop.mac_address)
    return " | ".join(parts)


def _format_service_date(value) -> str:
    return value.strftime("%d/%m/%Y")


def _format_service_datetime(value) -> str:
    return value.strftime("%Y-%m-%d %H:%M")

def generate_live_asset_report(db: Session, entity_type: str, asset) -> bytes:
    from io import BytesIO
    from openpyxl import load_workbook
    from ..models import ServiceJob

    # 1. Resolve project and template
    if entity_type == "meter":
        project = asset.loop.panel.project
    elif entity_type == "loop":
        project = asset.panel.project
    elif entity_type == "panel":
        project = asset.project
    else:
        raise ValueError("Invalid entity type")

    template_path = _resolve_template_path_for_project(project)
    if not template_path or not template_path.exists():
        raise FileNotFoundError("Template not found")

    workbook = load_workbook(template_path)

    # 2. Get latest service job for this asset to show "real-time" data
    latest_job = None
    if entity_type == "meter":
        latest_job = db.query(ServiceJob).filter(ServiceJob.meter_id == asset.id).order_by(ServiceJob.service_date.desc()).first()
    # For loop/panel we could find the latest job for any meter in them, but let's keep it simple for now

    # 3. Create a clean report workbook by keeping only necessary sheets
    # For a standalone asset report, we might want a cover and then the detail pages
    keep_sheets = ["Cover", "Page (1)", "Page (4)", "Page (5)"]
    for sheet_name in list(workbook.sheetnames):
        if sheet_name not in keep_sheets:
            del workbook[sheet_name]

    # 4. Populate the data
    report_number = f"LIVE-{entity_type.upper()}-{asset.id}"
    service_date = latest_job.service_date if latest_job else datetime.now()
    
    # Populate Cover
    cover = workbook["Cover"]
    cover["I21"] = project.name
    cover["I22"] = _format_service_date(service_date)
    cover["I23"] = f"LIVE {entity_type.upper()} REPORT"
    cover["I24"] = project.customer.name

    # Populate Page (1) - Intro
    intro = workbook["Page (1)"]
    _populate_common_template_header(intro, project.name, project.customer.name, report_number, service_date)
    intro["C12"] = f"This is a live real-time report for {entity_type} {getattr(asset, 'meter_code', getattr(asset, 'loop_code', getattr(asset, 'panel_code', '')))}."
    
    # Populate Detail Pages (4 & 5)
    if entity_type == "meter":
        meter = asset
        loop = meter.loop
        panel = loop.panel
        
        checklist_rows = []
        photos = []
        if latest_job:
            checklist_rows = _normalize_checklist_rows(latest_job)
            photos = latest_job.photos
        
        _populate_detail_sheet(
            sheet=workbook["Page (4)"],
            service_job=latest_job or ServiceJob(project=project, meter=meter, service_date=service_date, status="Pending"),
            panel=panel,
            loop=loop,
            meter=meter,
            customer_name=project.customer.name,
            project_name=project.name,
            report_number=report_number,
            service_date=service_date,
            page_number=3,
            total_pages=4,
            checklist_rows=checklist_rows[:10],
            photos=photos[:3],
            overview_note=latest_job.note if latest_job else "Live preview - No recent service history found.",
        )
        
        _populate_detail_sheet(
            sheet=workbook["Page (5)"],
            service_job=latest_job or ServiceJob(project=project, meter=meter, service_date=service_date, status="Pending"),
            panel=panel,
            loop=loop,
            meter=meter,
            customer_name=project.customer.name,
            project_name=project.name,
            report_number=report_number,
            service_date=service_date,
            page_number=4,
            total_pages=4,
            checklist_rows=checklist_rows[10:20],
            photos=photos[3:6],
            overview_note=latest_job.note if latest_job else "",
        )
    else:
        # For loop/panel, we just show basic info for now or we could list all meters.
        # But the user specifically asked for "specific item" report.
        pass

    # 5. Return as bytes
    out = BytesIO()
    workbook.save(out)
    return out.getvalue()

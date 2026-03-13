from datetime import datetime
from pathlib import Path
import re
from typing import Optional

from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from ..config import DEFAULT_PROJECT_WORKBOOK_PATH, UPLOADS_DIR
from ..models import Project
from ..utils.helpers import build_public_file_path
from ..utils.project_files import project_upload_dir

OAKWOOD_STATUS_ROWS = range(11, 43)
OAKWOOD_MODEL_COLUMNS = ("K", "L", "M", "N")
OAKWOOD_CLEAR_COLUMNS = ("B", "C", "D", "F", "G", "H", "O", "P", *OAKWOOD_MODEL_COLUMNS)


def export_project_assets_to_excel(db: Session, project: Project) -> dict:
    template_path, template_source = _resolve_template_path(project)
    workbook = load_workbook(template_path) if template_path else Workbook()

    if _is_oakwood_loop_workbook(workbook):
        _write_oakwood_loop_sheets(workbook, project)

    _write_standard_asset_sheets(workbook, project, template_source)

    if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
        sheet = workbook["Sheet"]
        if sheet.max_row <= 1 and sheet.max_column <= 1 and sheet["A1"].value in (None, ""):
            workbook.remove(sheet)

    output_dir = project_upload_dir(UPLOADS_DIR, project, "documents", "asset_exports")
    output_dir.mkdir(parents=True, exist_ok=True)

    exported_at = datetime.now()
    timestamp = exported_at.strftime("%Y-%m-%d_%H%M%S")
    base_name = _export_base_name(project, template_path)
    output_path = output_dir / f"{base_name}_{timestamp}.xlsx"
    workbook.save(output_path)

    return {
        "file_name": output_path.name,
        "file_path": build_public_file_path(UPLOADS_DIR, output_path),
        "saved_to": output_path.relative_to(UPLOADS_DIR.parent).as_posix(),
        "template_source": template_source,
        "exported_at": exported_at.isoformat(timespec="seconds"),
    }


def _resolve_template_path(project: Project) -> tuple[Optional[Path], str]:
    if project.project_workbook_file_path:
        candidate = UPLOADS_DIR / project.project_workbook_file_path.removeprefix("/uploads/")
        if candidate.exists():
            return candidate, "project_workbook"

    if project.template_file_path:
        candidate = UPLOADS_DIR / project.template_file_path.removeprefix("/uploads/")
        if candidate.exists() and _is_oakwood_loop_workbook(load_workbook(candidate, read_only=True)):
            return candidate, "legacy_project_workbook"

    if DEFAULT_PROJECT_WORKBOOK_PATH.exists():
        return DEFAULT_PROJECT_WORKBOOK_PATH, "default_workbook"

    return None, "generated"


def _export_base_name(project: Project, template_path: Optional[Path]) -> str:
    if template_path is not None:
        source_name = template_path.stem
        if _looks_like_generated_upload_name(source_name):
            source_name = f"{project.name} assets"
    else:
        source_name = f"{project.name} assets"
    cleaned = re.sub(r'[<>:"/\\\\|?*]+', "-", str(source_name)).strip().strip(".")
    return cleaned or f"project-{project.id}-assets"


def _looks_like_generated_upload_name(value: str) -> bool:
    return bool(re.fullmatch(r"FILE-\d{14}-[A-F0-9]{6}", value.strip()))


def _collect_asset_rows(project: Project) -> list[dict]:
    rows = []
    panels = sorted(project.panels, key=lambda item: ((item.panel_code or "").lower(), item.id))
    for panel in panels:
        loops = sorted(panel.loops, key=lambda item: ((item.loop_code or "").lower(), item.id))
        for loop in loops:
            meters = sorted(loop.meters, key=lambda item: ((item.meter_code or "").lower(), item.id))
            if not meters:
                rows.append(
                    {
                        "panel": panel,
                        "loop": loop,
                        "meter": None,
                    }
                )
                continue
            for meter in meters:
                rows.append({"panel": panel, "loop": loop, "meter": meter})
    return rows


def _write_standard_asset_sheets(workbook, project: Project, template_source: str):
    rows = _collect_asset_rows(project)
    exported_at = datetime.now().isoformat(timespec="seconds")

    metadata_sheet = _replace_sheet(workbook, "EMS_METADATA")
    _append_rows(
        metadata_sheet,
        [
            ("field", "value"),
            ("project_id", project.id),
            ("project_name", project.name),
            ("location", project.location),
            ("customer_name", project.customer.name if project.customer else ""),
            ("template_source", template_source),
            ("report_template_file_path", project.template_file_path or ""),
            ("project_workbook_file_path", project.project_workbook_file_path or ""),
            ("exported_at", exported_at),
            ("note", "Save-as workbook generated from current Asset Manager data."),
        ],
    )

    panels_sheet = _replace_sheet(workbook, "EMS_PANELS")
    panels_sheet.append(["panel_id", "panel_code", "panel_name", "serial_number", "location_note"])
    for panel in sorted(project.panels, key=lambda item: ((item.panel_code or "").lower(), item.id)):
        panels_sheet.append([
            panel.id,
            panel.panel_code,
            panel.panel_name,
            panel.serial_number or "",
            panel.location_note or "",
        ])

    loops_sheet = _replace_sheet(workbook, "EMS_LOOPS")
    loops_sheet.append([
        "loop_id",
        "panel_id",
        "panel_code",
        "loop_code",
        "loop_name",
        "converter_name",
        "converter_ip",
        "mac_address",
    ])
    for panel in sorted(project.panels, key=lambda item: ((item.panel_code or "").lower(), item.id)):
        for loop in sorted(panel.loops, key=lambda item: ((item.loop_code or "").lower(), item.id)):
            loops_sheet.append([
                loop.id,
                panel.id,
                panel.panel_code,
                loop.loop_code,
                loop.loop_name,
                loop.converter_name or "",
                loop.converter_ip or "",
                loop.mac_address or "",
            ])

    meters_sheet = _replace_sheet(workbook, "EMS_METERS")
    meters_sheet.append([
        "meter_id",
        "loop_id",
        "panel_code",
        "loop_code",
        "meter_code",
        "meter_name",
        "serial_number",
        "device_address",
        "model",
        "ct_ratio",
        "baud_rate",
        "status",
    ])
    for row in rows:
        meter = row["meter"]
        if meter is None:
            continue
        loop = row["loop"]
        panel = row["panel"]
        meters_sheet.append([
            meter.id,
            loop.id,
            panel.panel_code,
            loop.loop_code,
            meter.meter_code,
            meter.meter_name,
            meter.serial_number or "",
            meter.device_address or "",
            meter.model or "",
            meter.ct_ratio or "",
            meter.baud_rate or "",
            meter.status or "",
        ])

    assets_sheet = _replace_sheet(workbook, "EMS_ASSETS")
    assets_sheet.append([
        "panel_code",
        "panel_name",
        "panel_serial_number",
        "panel_location_note",
        "loop_code",
        "loop_name",
        "converter_name",
        "converter_ip",
        "mac_address",
        "meter_code",
        "meter_name",
        "meter_serial_number",
        "device_address",
        "model",
        "ct_ratio",
        "baud_rate",
        "status",
    ])
    for row in rows:
        loop = row["loop"]
        panel = row["panel"]
        meter = row["meter"]
        assets_sheet.append([
            panel.panel_code,
            panel.panel_name,
            panel.serial_number or "",
            panel.location_note or "",
            loop.loop_code,
            loop.loop_name,
            loop.converter_name or "",
            loop.converter_ip or "",
            loop.mac_address or "",
            meter.meter_code if meter else "",
            meter.meter_name if meter else "",
            meter.serial_number if meter else "",
            meter.device_address if meter else "",
            meter.model if meter else "",
            meter.ct_ratio if meter else "",
            meter.baud_rate if meter else "",
            meter.status if meter else "",
        ])

    for sheet_name in ("EMS_METADATA", "EMS_PANELS", "EMS_LOOPS", "EMS_METERS", "EMS_ASSETS"):
        workbook[sheet_name].freeze_panes = "A2"


def _replace_sheet(workbook, title: str):
    if title in workbook.sheetnames:
        workbook.remove(workbook[title])
    return workbook.create_sheet(title)


def _append_rows(sheet, rows):
    for row in rows:
        sheet.append(list(row))


def _is_oakwood_loop_workbook(workbook) -> bool:
    sheet_names = set(workbook.sheetnames)
    return "Cover" in sheet_names and any(name.lower().startswith("loop") for name in workbook.sheetnames)


def _write_oakwood_loop_sheets(workbook, project: Project):
    loop_entries = []
    for panel in sorted(project.panels, key=lambda item: ((item.panel_code or "").lower(), item.id)):
        for loop in sorted(panel.loops, key=lambda item: ((item.loop_code or "").lower(), item.id)):
            meters = sorted(loop.meters, key=lambda item: ((item.meter_code or "").lower(), item.id))
            loop_entries.append({"panel": panel, "loop": loop, "meters": meters})

    loop_sheets = [workbook[name] for name in workbook.sheetnames if name.lower().startswith("loop")]
    if not loop_sheets:
        return

    template_sheet = loop_sheets[0]
    assigned_sheets = []
    for index, entry in enumerate(loop_entries):
        if index < len(loop_sheets):
            sheet = loop_sheets[index]
        else:
            sheet = workbook.copy_worksheet(template_sheet)
        desired_title = _unique_sheet_title(workbook, entry["loop"].loop_code or f"Loop{index + 1}", current_sheet=sheet)
        sheet.title = desired_title
        _populate_oakwood_loop_sheet(sheet, entry["panel"], entry["loop"], entry["meters"])
        assigned_sheets.append(sheet)

    for sheet in loop_sheets[len(loop_entries):]:
        _populate_oakwood_loop_sheet(sheet, None, None, [])

    if not loop_entries:
        _populate_oakwood_loop_sheet(template_sheet, None, None, [])


def _populate_oakwood_loop_sheet(sheet, panel, loop, meters):
    sheet["O3"] = panel.panel_name if panel else "No panel assigned"
    sheet["T3"] = loop.loop_name if loop else "No loop assigned"
    sheet["C4"] = loop.converter_name if loop and loop.converter_name else ""
    sheet["C5"] = loop.converter_ip if loop and loop.converter_ip else ""
    sheet["H5"] = loop.mac_address if loop and loop.mac_address else ""

    distinct_models = []
    for meter in meters:
        model = str(meter.model or "").strip()
        if model and model not in distinct_models:
            distinct_models.append(model)
    for column, model in zip(OAKWOOD_MODEL_COLUMNS, distinct_models[: len(OAKWOOD_MODEL_COLUMNS)]):
        sheet[f"{column}10"] = model
    for column in OAKWOOD_MODEL_COLUMNS[len(distinct_models[: len(OAKWOOD_MODEL_COLUMNS)]):]:
        sheet[f"{column}10"] = ""

    model_headers = {column: str(sheet[f"{column}10"].value or "").strip().lower() for column in OAKWOOD_MODEL_COLUMNS}

    for row_index in OAKWOOD_STATUS_ROWS:
        for column in OAKWOOD_CLEAR_COLUMNS:
            sheet[f"{column}{row_index}"] = ""

    for row_index, meter in zip(OAKWOOD_STATUS_ROWS, meters[: len(OAKWOOD_STATUS_ROWS)]):
        sheet[f"B{row_index}"] = meter.meter_code or ""
        sheet[f"C{row_index}"] = meter.serial_number or ""
        sheet[f"D{row_index}"] = panel.panel_code if panel else ""
        sheet[f"F{row_index}"] = meter.device_address or ""
        sheet[f"G{row_index}"] = meter.baud_rate or ""
        sheet[f"H{row_index}"] = meter.ct_ratio or ""

        meter_model = str(meter.model or "").strip().lower()
        for column in OAKWOOD_MODEL_COLUMNS:
            sheet[f"{column}{row_index}"] = "P" if meter_model and model_headers.get(column) == meter_model else ""

        status = str(meter.status or "").strip().lower()
        sheet[f"O{row_index}"] = "P" if status in {"active", "maintenance"} else ""
        sheet[f"P{row_index}"] = "P" if status == "offline" else ""


def _unique_sheet_title(workbook, desired_title: str, current_sheet=None) -> str:
    clean_title = re.sub(r"[:\\\\/?*\\[\\]]+", "-", str(desired_title or "Loop")).strip() or "Loop"
    clean_title = clean_title[:31]
    existing_titles = {sheet.title for sheet in workbook.worksheets if sheet is not current_sheet}
    if clean_title not in existing_titles:
        return clean_title

    base_title = clean_title[:28] or "Loop"
    index = 2
    while True:
        candidate = f"{base_title}-{index}"[:31]
        if candidate not in existing_titles:
            return candidate
        index += 1

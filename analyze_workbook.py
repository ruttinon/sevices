import openpyxl
from openpyxl import load_workbook
import os

# Analyze workbook structure
wb_path = 'uploads/projects/Oakwood Suites Bangkok/Oakwood RMS Maintenance/documents/workbooks/Maintenance Agreement Format - Oakwood.xlsx'
wb = load_workbook(wb_path, read_only=True)

print('=== WORKBOOK SHEETS ===')
for name in wb.sheetnames:
    print(f'  - {name}')

print()
print('=== Analyzing Cover Sheet ===')
sheet = wb['Cover']
for row in range(1, min(20, sheet.max_row + 1)):
    values = [str(cell.value)[:40] if cell.value else '' for cell in sheet[row]]
    if any(values):
        print(f'  Row {row}: {values[:6]}')

print()
print('=== Analyzing Loop Sheets ===')
for sheet_name in wb.sheetnames:
    if sheet_name.startswith('Loop'):
        sheet = wb[sheet_name]
        print(f'\n--- {sheet_name} ---')
        for row in range(1, min(15, sheet.max_row + 1)):
            values = [str(cell.value)[:40] if cell.value else '' for cell in sheet[row]]
            if any(v for v in values if v):
                print(f'  Row {row}: {values[:6]}')

wb.close()
